from binance.client import Client
from binance import ThreadedWebsocketManager
import pandas as pd
import numpy as np
import time
from datetime import datetime
import requests
import os
import json
import argparse
import openpyxl
import math
from db import init_db, insert_trade, update_strategy_status
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from threading import Thread, Lock
from queue import Queue

# ===== ЗАГРУЗКА КОНФИГА =====
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=True)
args = parser.parse_args()

with open(args.config, "r") as f:
    config = json.load(f)

BOT_NAME = config["NAME"]

load_dotenv()

# Фильтр по объёму свечи (quote volume текущей свечи, не 24h)
USE_VOLUME_FILTER  = config.get("USE_VOLUME_FILTER", False)
VOLUME_MIN_BUY     = float(config.get("VOLUME_MIN_BUY",  0))
VOLUME_MAX_BUY     = float(config.get("VOLUME_MAX_BUY",  0))
VOLUME_MIN_SELL    = float(config.get("VOLUME_MIN_SELL", 0))
VOLUME_MAX_SELL    = float(config.get("VOLUME_MAX_SELL", 0))

# Фильтр по NATR
USE_NATR_FILTER    = config.get("USE_NATR_FILTER", False)
NATR_MIN_BUY       = float(config.get("NATR_MIN_BUY",  0))
NATR_MAX_BUY       = float(config.get("NATR_MAX_BUY",  0))
NATR_MIN_SELL      = float(config.get("NATR_MIN_SELL", 0))
NATR_MAX_SELL      = float(config.get("NATR_MAX_SELL", 0))

# Фильтр по корреляции с BTC
USE_CORREL_FILTER  = config.get("USE_CORREL_FILTER", False)
CORREL_MIN_BUY     = float(config.get("CORREL_MIN_BUY",  0))
CORREL_MAX_BUY     = float(config.get("CORREL_MAX_BUY",  0))
CORREL_MIN_SELL    = float(config.get("CORREL_MIN_SELL", 0))
CORREL_MAX_SELL    = float(config.get("CORREL_MAX_SELL", 0))

# Дни недели когда не торгуем
SKIP_DAYS          = [d.lower() for d in config.get("SKIP_DAYS", [])]
# Пример: ["monday", "saturday", "sunday"]

# ================= НАСТРОЙКИ =================
MIN_24H_VOLUME   = config["MIN_24H_VOLUME"]
LOOKBACK_CANDLES = config["LOOKBACK_CANDLES"]
BTC_LOOKBACK     = config["BTC_LOOKBACK"]
ATR_LEN          = config["ATR_LEN"]

EMA_FAST         = config["EMA_FAST"]
EMA_SLOW         = config["EMA_SLOW"]

# Swing Detection Lookback
SWING_LENGTH     = config["SWING_LENGTH"]  # 24, 50, 100

# Slope метод — только один true
USE_ATR_SLOPE    = config.get("USE_ATR_SLOPE", True)
USE_STDEV_SLOPE  = config.get("USE_STDEV_SLOPE", False)
USE_LINREG_SLOPE = config.get("USE_LINREG_SLOPE", False)
SLOPE_MULT       = float(config.get("SLOPE_MULT", 1.0))

# EMA фильтр
USE_EMA_FILTER   = config.get("USE_EMA_FILTER", True)

# RelVol фильтр
USE_RELVOL_FILTER = config.get("USE_RELVOL_FILTER", False)
RELVOL_ANCHOR     = config.get("RELVOL_ANCHOR", "1d")
RELVOL_LENGTH     = config.get("RELVOL_LENGTH", 5)
RELVOL_MIN_BUY    = float(config.get("RELVOL_MIN_BUY",  0.0))
RELVOL_MAX_BUY    = float(config.get("RELVOL_MAX_BUY",  0.0))
RELVOL_MIN_SELL   = float(config.get("RELVOL_MIN_SELL", 0.0))
RELVOL_MAX_SELL   = float(config.get("RELVOL_MAX_SELL", 0.0))

ANCHOR_MAP = {
    "1h": Client.KLINE_INTERVAL_1HOUR,
    "4h": Client.KLINE_INTERVAL_4HOUR,
    "1d": Client.KLINE_INTERVAL_1DAY,
    "1w": Client.KLINE_INTERVAL_1WEEK,
}

EXCEL_STRAT_START_COL = 14  # колонка N

CHAT_ID   = os.getenv("CHAT_ID")
BOT_TOKEN = os.getenv("BOT_TOKEN")

client = Client()
BLACKLIST = {
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT"
}

# ================= TRADES =================
TRADE_STATE_FILE   = f"trades_state_{BOT_NAME}.json"
EXCEL_FILE         = f"trades_{BOT_NAME}.xlsx"
ACTIVE_TRADES_FILE = f"active_trades_{BOT_NAME}.json"

TRADES_LOCK = Lock()
EXCEL_LOCK  = Lock()
_ID_LOCK    = Lock()

SHEET_MAP = {
    "CONFTL3":  "conftl3",
    "CONFTL5":  "conftl5",
    "CONFTL7":  "conftl7",
    "CONFTL9":  "conftl9",
    "CONFTL11": "conftl11",
}

STRATEGIES = {
    "3:1":  {"tp": 0.03,  "sl": -0.01},
    "6:1":  {"tp": 0.06,  "sl": -0.01},
    "6:2":  {"tp": 0.06,  "sl": -0.02},
    "10:3": {"tp": 0.10,  "sl": -0.03},
    "12:4": {"tp": 0.12,  "sl": -0.04},
}

def check_and_close_strategies(symbol, price_high, price_low):
    closed_trades = []
    with TRADES_LOCK:
        for trade_id, trade in list(ACTIVE_TRADES.items()):
            if trade["symbol"] != symbol:
                continue
            for strat_name, strat in trade["strategies"].items():
                if strat["status"] != "OPEN":
                    continue
                result = None
                if trade["side"] == "BUY":
                    if price_low <= strat["sl"]:
                        result = "SL"
                    elif price_high >= strat["tp"]:
                        result = "TP"
                else:
                    if price_high >= strat["sl"]:
                        result = "SL"
                    elif price_low <= strat["tp"]:
                        result = "TP"
                if result:
                    strat["status"] = result
                    strat["close_time"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                    update_strategy_status(trade_id, strat_name, result)
                    Thread(
                        target=update_trade_status_in_excel,
                        args=(trade_id, strat_name, result),
                        daemon=True
                    ).start()
            if all(s["status"] != "OPEN" for s in trade["strategies"].values()):
                closed_trades.append(trade_id)
        for tid in closed_trades:
            del ACTIVE_TRADES[tid]
    if closed_trades:
        save_active_trades()
    return closed_trades

def load_trade_id():
    if not os.path.exists(TRADE_STATE_FILE):
        return 0
    with open(TRADE_STATE_FILE, "r") as f:
        return json.load(f).get("last_trade_id", 0)

def save_trade_id(tid):
    with open(TRADE_STATE_FILE, "w") as f:
        json.dump({"last_trade_id": tid}, f)

def save_active_trades():
    with TRADES_LOCK:
        with open(ACTIVE_TRADES_FILE, "w") as f:
            json.dump(ACTIVE_TRADES, f)

def load_active_trades():
    if not os.path.exists(ACTIVE_TRADES_FILE):
        return {}
    with open(ACTIVE_TRADES_FILE, "r") as f:
        return json.load(f)

ACTIVE_TRADES = load_active_trades()
LAST_TRADE_ID = load_trade_id()

def get_next_trade_id():
    global LAST_TRADE_ID
    with _ID_LOCK:
        LAST_TRADE_ID += 1
        save_trade_id(LAST_TRADE_ID)
        return f"{LAST_TRADE_ID:05d}"

# ================= TELEGRAM =================
def send_telegram(message: str):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": message}
    try:
        requests.post(url, data=payload, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")

# ================= EXCEL =================
def write_trade_to_excel(trade_id, trade_info, vol_text, vol24, corr_text):
    sheet_name = SHEET_MAP.get(BOT_NAME, "conftl1")

    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            for sn in SHEET_MAP.values():
                if sn not in wb.sheetnames:
                    wb.create_sheet(sn)
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        headers = {
            "I":"Дата","B":"Время","C":"День","D":"Тикет","E":"Объем",
            "F":"Trade_id","G":"Тип","H":"Импульс","J":"Цена входа",
            "L":"Корреляция","M":"NATR%",
            "N":"3:1","O":"6:1","P":"6:2","Q":"10:3","R":"12:4",
            "S":"ATR Ratio 50","T":"ATR Ratio 75","U":"ATR Ratio 100",
            "V":"Expansion 3","W":"Expansion 5",
            "X":"Свинг"
        }
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            for col, header in headers.items():
                ws[f"{col}1"] = header

        next_row = ws.max_row + 1
        dt = datetime.now()
        ws["I"+str(next_row)] = dt.strftime("%d.%m.%Y")
        ws["B"+str(next_row)] = dt.strftime("%H:%M:%S")
        ws["C"+str(next_row)] = dt.strftime("%a")
        ws["D"+str(next_row)] = trade_info["symbol"]
        ws["E"+str(next_row)] = vol24
        ws["F"+str(next_row)] = trade_id
        ws["G"+str(next_row)] = ", ".join(trade_info["signals"])
        ws["H"+str(next_row)] = vol_text
        ws["J"+str(next_row)] = trade_info["entry_price"]
        ws["L"+str(next_row)] = corr_text
        ws["M"+str(next_row)] = trade_info["natr"]
        ws["S"+str(next_row)] = trade_info["atr_ratio_50"]
        ws["T"+str(next_row)] = trade_info["atr_ratio_75"]
        ws["U"+str(next_row)] = trade_info["atr_ratio_100"]
        ws["V"+str(next_row)] = trade_info["expansion_3"]
        ws["W"+str(next_row)] = trade_info["expansion_5"]
        ws["X"+str(next_row)] = trade_info["swing_num"]

        for idx, s in enumerate(STRATEGIES.keys()):
            col = get_column_letter(EXCEL_STRAT_START_COL + idx)
            ws[f"{col}{next_row}"] = trade_info["strategies"][s]["status"]

        wb.active = wb[sheet_name]
        wb.save(EXCEL_FILE)

def update_trade_status_in_excel(trade_id, strategy_name, status):
    sheet_name = SHEET_MAP.get(BOT_NAME, "conftl1")

    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]

        for row in range(2, ws.max_row+1):
            if str(ws[f"F{row}"].value) == trade_id:
                col_map_status  = {"3:1":"N","6:1":"O","6:2":"P","10:3":"Q","12:4":"R"}
                col_s = col_map_status[strategy_name]
                ws[f"{col_s}{row}"] = status
                break

        wb.active = wb[sheet_name]
        wb.save(EXCEL_FILE)

# ================= INDICATORS =================
def calculate_atr(df, period):
    hl = df["high"] - df["low"]
    hc = (df["high"] - df["close"].shift()).abs()
    lc = (df["low"] - df["close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def calculate_atr_ratio(df, avg_len, atr_len=14):
    atr = calculate_atr(df, atr_len)
    atr_avg = atr.rolling(avg_len).mean()

    if pd.isna(atr_avg.iloc[-2]) or atr_avg.iloc[-2] == 0:
        return 0

    ratio = atr.iloc[-2] / atr_avg.iloc[-2]
    return round(ratio, 4)


def calculate_expansion(df, lookback, natr_len=14):
    atr = calculate_atr(df, natr_len)

    last_close = df["close"].iloc[-2]
    prev_close = df["close"].iloc[-2 - lookback]

    atr_value = atr.iloc[-2]

    if prev_close == 0 or last_close == 0 or atr_value == 0:
        return 0

    # Pine:
    # natr = atr / close * 100
    natr = (atr_value / last_close) * 100

    # Pine:
    # abs(close - close[lookback]) / close[lookback] * 100
    price_change_pct = abs(last_close - prev_close) / prev_close * 100

    # Pine:
    # expansion = priceChangePct / (natr * sqrt(lookback))
    expansion = price_change_pct / (natr * math.sqrt(lookback))

    return round(expansion, 4)

def get_liquid_futures_symbols():
    tickers = client._request_futures_api(method="get", path="ticker/24hr")
    symbols = []
    for t in tickers:
        symbol = t["symbol"]
        if not symbol.endswith("USDT") or symbol in BLACKLIST:
            continue
        if float(t["quoteVolume"]) < MIN_24H_VOLUME:
            continue
        symbols.append(symbol)
    return symbols

def get_symbols_with_open_trades():
    with TRADES_LOCK:
        return {
            trade["symbol"] for trade in ACTIVE_TRADES.values()
            if any(s["status"] == "OPEN" for s in trade["strategies"].values())
        }

def get_btc_returns():
    try:
        klines_btc = client.futures_klines(
            symbol="BTCUSDT", interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
        )
        df_btc = pd.DataFrame(klines_btc, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_btc["close"] = df_btc["close"].astype(float)
        return df_btc["close"].pct_change()
    except Exception as e:
        print(f"Ошибка загрузки BTC свечей: {e}")
        return None

def get_swing_num(df, side, n=5):
    """Наидальнейшая свеча из n предыдущих с low ниже (BUY) или high выше (SELL)."""
    current = df.iloc[-2]
    result = 0
    for i in range(1, n + 1):
        idx = -2 - i
        if abs(idx) > len(df):
            break
        candle = df.iloc[idx]
        if side == "BUY" and candle["low"] < current["low"]:
            result = i
        if side == "SELL" and candle["high"] > current["high"]:
            result = i
    return result

def calculate_relvol(symbol, current_vol, anchor_interval, length):
    try:
        anchor_limit = length * 24 + 5 if anchor_interval == Client.KLINE_INTERVAL_1HOUR else length + 5
        klines = client.futures_klines(
            symbol=symbol, interval=anchor_interval, limit=anchor_limit
        )
        df_anchor = pd.DataFrame(klines, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_anchor["open_time"] = pd.to_datetime(df_anchor["open_time"], unit="ms")
        df_anchor["quote_volume"] = df_anchor["close"].astype(float) * df_anchor["volume"].astype(float)

        current_hour = df_anchor["open_time"].iloc[-2].hour

        if anchor_interval == Client.KLINE_INTERVAL_1HOUR:
            same_hour = df_anchor[df_anchor["open_time"].dt.hour == current_hour].iloc[-(length+1):-1]
        else:
            same_hour = df_anchor.iloc[-(length+1):-1]

        if len(same_hour) == 0:
            return None

        avg_vol = same_hour["quote_volume"].mean()
        if avg_vol == 0:
            return None

        return round(current_vol / avg_vol, 3)
    except Exception as e:
        print(f"Ошибка RelVol {symbol}: {e}")
        return None

def find_pivot_high(df, length, idx):
    """Проверяет является ли свеча idx свинг хаем."""
    if idx < length or idx >= len(df) - length:
        return False
    center = df["high"].iloc[idx]
    for i in range(idx - length, idx + length + 1):
        if i == idx:
            continue
        if df["high"].iloc[i] >= center:
            return False
    return True

def find_pivot_low(df, length, idx):
    """Проверяет является ли свеча idx свинг лоу."""
    if idx < length or idx >= len(df) - length:
        return False
    center = df["low"].iloc[idx]
    for i in range(idx - length, idx + length + 1):
        if i == idx:
            continue
        if df["low"].iloc[i] <= center:
            return False
    return True

def calculate_slope(df, length):
    """Рассчитывает наклон по выбранному методу."""
    close = df["close"]
    if USE_ATR_SLOPE:
        atr = calculate_atr(df, length)
        return (atr / length * SLOPE_MULT).iloc[-2]
    elif USE_STDEV_SLOPE:
        stdev = close.rolling(length).std()
        return (stdev / length * SLOPE_MULT).iloc[-2]
    elif USE_LINREG_SLOPE:
        n = len(df)
        idx = pd.Series(range(n))
        sma_cn = (close * idx).rolling(length).mean()
        sma_c  = close.rolling(length).mean()
        sma_n  = idx.rolling(length).mean()
        var_n  = idx.rolling(length).var()
        slope  = (sma_cn - sma_c * sma_n) / var_n / 2 * SLOPE_MULT
        return abs(slope.iloc[-2])
    return 0.0

def apply_range_filter(value, min_val, max_val):
    """
    Проверяет попадает ли value в диапазон [min_val, max_val].
    0 = фильтр по этой границе выключен.
    """
    if value is None:
        return False
    if min_val > 0 and value < min_val:
        return False
    if max_val > 0 and value > max_val:
        return False
    return True

def check_trendline_signal(symbol):
    klines = client.futures_klines(
        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=LOOKBACK_CANDLES
    )
    df = pd.DataFrame(klines, columns=[
        "open_time","open","high","low","close",
        "volume","close_time","quote_volume",
        "trades","taker_buy_base","taker_buy_quote","ignore"
    ])
    for c in ["open","high","low","close","volume"]:
        df[c] = df[c].astype(float)

    df["ema20"]  = df["close"].ewm(span=EMA_FAST, adjust=False).mean()
    df["ema200"] = df["close"].ewm(span=EMA_SLOW, adjust=False).mean()
    df["atr"]    = calculate_atr(df, ATR_LEN)
    df["natr"]   = (df["atr"] / df["close"]) * 100
    df["quote_volume"] = df["close"] * df["volume"]

    atr_ratio_50  = calculate_atr_ratio(df, 50)
    atr_ratio_75  = calculate_atr_ratio(df, 75)
    atr_ratio_100 = calculate_atr_ratio(df, 100)

    expansion_3 = calculate_expansion(df, 3)
    expansion_5 = calculate_expansion(df, 5)

    last_idx = len(df) - 2  # последняя закрытая свеча
    last = df.iloc[last_idx]
    prev = df.iloc[last_idx - 1]

    # ===== EMA фильтр =====
    bull_trend = last["ema20"] > last["ema200"]
    bear_trend = last["ema20"] < last["ema200"]
    ema_bull_ok = bull_trend if USE_EMA_FILTER else True
    ema_bear_ok = bear_trend if USE_EMA_FILTER else True

    # ===== Рассчитываем наклон =====
    slope = calculate_slope(df, SWING_LENGTH)

    # ===== Ищем последний свинг хай и строим нисходящий трендлайн =====
    upper = None
    slope_ph = slope
    for i in range(last_idx - 1, max(last_idx - LOOKBACK_CANDLES, SWING_LENGTH), -1):
        if find_pivot_high(df, SWING_LENGTH, i):
            upper = df["high"].iloc[i]
            # Экстраполируем трендлайн до текущей свечи
            bars_since = last_idx - i
            upper = upper - slope_ph * bars_since
            break

    # ===== Ищем последний свинг лоу и строим восходящий трендлайн =====
    lower = None
    slope_pl = slope
    for i in range(last_idx - 1, max(last_idx - LOOKBACK_CANDLES, SWING_LENGTH), -1):
        if find_pivot_low(df, SWING_LENGTH, i):
            lower = df["low"].iloc[i]
            bars_since = last_idx - i
            lower = lower + slope_pl * bars_since
            break

    if upper is None or lower is None:
        return None

    # Предыдущие значения трендлайнов
    upper_prev = upper + slope_ph  # на одну свечу назад
    lower_prev = lower - slope_pl

    # ===== Пробои =====
    # BUY: цена пробивает нисходящий трендлайн вверх
    buy_breakout  = prev["close"] <= upper_prev and last["close"] > upper
    # SELL: цена пробивает восходящий трендлайн вниз
    sell_breakout = prev["close"] >= lower_prev and last["close"] < lower

    # ===== Volume фильтр (quote volume текущей свечи в млн) =====
    candle_vol_m = last["quote_volume"] / 1_000_000
    if USE_VOLUME_FILTER:
        vol_buy_ok  = apply_range_filter(candle_vol_m, VOLUME_MIN_BUY,  VOLUME_MAX_BUY)
        vol_sell_ok = apply_range_filter(candle_vol_m, VOLUME_MIN_SELL, VOLUME_MAX_SELL)
    else:
        vol_buy_ok  = True
        vol_sell_ok = True

    # ===== NATR фильтр =====
    natr_val = last["natr"]
    if USE_NATR_FILTER:
        natr_buy_ok  = apply_range_filter(natr_val, NATR_MIN_BUY,  NATR_MAX_BUY)
        natr_sell_ok = apply_range_filter(natr_val, NATR_MIN_SELL, NATR_MAX_SELL)
    else:
        natr_buy_ok  = True
        natr_sell_ok = True

    signals = []
    if buy_breakout and ema_bull_ok and vol_buy_ok and natr_buy_ok:
        signals.append("BUY_TREND")
        
    if sell_breakout and ema_bear_ok vol_sell_ok and natr_sell_ok:
        signals.append("SELL_TREND")

    if not signals:
        return None

    # ===== RelVol =====
    anchor_interval = ANCHOR_MAP.get(RELVOL_ANCHOR, Client.KLINE_INTERVAL_1DAY)
    current_vol = last["quote_volume"]
    relvol = calculate_relvol(symbol, current_vol, anchor_interval, RELVOL_LENGTH) if USE_RELVOL_FILTER else None

    if USE_RELVOL_FILTER and relvol is not None:
        side = "BUY" if any("BUY" in s for s in signals) else "SELL"
        if side == "BUY":
            rv_ok = (RELVOL_MIN_BUY  <= 0 or relvol >= RELVOL_MIN_BUY)  and (RELVOL_MAX_BUY  <= 0 or relvol <= RELVOL_MAX_BUY)
        else:
            rv_ok = (RELVOL_MIN_SELL <= 0 or relvol >= RELVOL_MIN_SELL) and (RELVOL_MAX_SELL <= 0 or relvol <= RELVOL_MAX_SELL)
        if not rv_ok:
            return None

    # ===== Свинг номер =====
    side_for_swing = "BUY" if any("BUY" in s for s in signals) else "SELL"
    swing_num = get_swing_num(df, side_for_swing, 5)

    # ===== Объём 24h =====
    ticker_24h = client.futures_ticker(symbol=symbol)
    volume_24h = float(ticker_24h["quoteVolume"])

    avg_vol = df["quote_volume"].iloc[-52:-2].mean()

    return {
        "symbol":    symbol,
        "signals":   signals,
        "close":     last["close"],
        "ema20":     last["ema20"],
        "ema200":    last["ema200"],
        "natr":      round(last["natr"], 3),
        "volText":   f"x{current_vol/avg_vol:.2f}" if avg_vol > 0 else "x0",
        "volume_24h": volume_24h,
        "swing_num": swing_num,
        "relvol":    relvol if relvol is not None else "N/A",
        "upper":     round(upper, 6),
        "lower":     round(lower, 6),
        "atr_ratio_50": atr_ratio_50,
        "atr_ratio_75": atr_ratio_75,
        "atr_ratio_100": atr_ratio_100,
        "expansion_3": expansion_3,
        "expansion_5": expansion_5,
    }

# ================= MAIN =================
def main():
    init_db()
    symbols = get_liquid_futures_symbols()
    print(f"✅ Ликвидные токены: {len(symbols)}")

    def update_symbols_periodically():
        nonlocal symbols
        while True:
            time.sleep(3600)
            try:
                liquid = get_liquid_futures_symbols()
                open_syms = get_symbols_with_open_trades()
                symbols = list(set(liquid) | open_syms)
                print(f"♻️ Обновление: {len(liquid)} ликвидных + {len(open_syms)} с открытыми позициями")
            except Exception as e:
                print(f"Ошибка обновления токенов: {e}")

    Thread(target=update_symbols_periodically, daemon=True).start()

    task_queue = Queue()

    def process_signal(msg):
    # Быстрая проверка дня — до любых вычислений
    today = datetime.utcnow().strftime("%A").lower()
    if today in SKIP_DAYS:
        return
        try:
            if msg.get("e") == "error":
                err_msg = msg.get("m","неизвестно")
                print(f"🔴 WebSocket ошибка: {msg}")
                if "reset" in str(err_msg).lower() or "closed" in str(err_msg).lower():
                    send_telegram(f"🔴 {BOT_NAME} WebSocket ошибка: {err_msg}")
                return

            if 'data' not in msg or 'k' not in msg['data']:
                return
            candle = msg['data']['k']
            symbol = candle['s']
            if not candle['x']:
                return

            price_high = float(candle["h"])
            price_low  = float(candle["l"])

            # ===== Закрытие открытых стратегий =====
            check_and_close_strategies(symbol, price_high, price_low)

            # Новые сигналы только для символов из основного списка
            if symbol not in symbols:
                return

            # ===== Новые сигналы =====
            res = check_trendline_signal(symbol)
            if not res:
                return

            entry_price = res["close"]
            side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"

            # ===== Корреляция BTC =====
            try:
                btc_returns = get_btc_returns()
                if btc_returns is not None:
                    klines_sym = client.futures_klines(
                        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
                    )
                    df_sym = pd.DataFrame(klines_sym, columns=[
                        "open_time","open","high","low","close","volume",
                        "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
                    ])
                    df_sym["close"] = df_sym["close"].astype(float)
                    symbol_returns = df_sym["close"].pct_change()
                    btc_subset = btc_returns[-len(symbol_returns):]
                    corr = btc_subset.corr(symbol_returns)
                    corr_text = round(float(corr), 2) if corr is not None else "N/A"
                else:
                    corr_text = "N/A"
            except Exception as e:
                print(f"Ошибка корреляции {symbol}: {e}")
                corr_text = "N/A"

            # После расчёта corr_text добавь:

            # ===== Фильтр по дням =====
            today = datetime.utcnow().strftime("%A").lower()  # "monday", "tuesday" ...
            if today in SKIP_DAYS:
                print(f"⏭️ {symbol} пропущен — день {today} в SKIP_DAYS")
                return

            # ===== Фильтр по корреляции =====
            if USE_CORREL_FILTER and isinstance(corr_text, float):
                side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"
                if side == "BUY":
                    if not apply_range_filter(corr_text, CORREL_MIN_BUY, CORREL_MAX_BUY):
                        print(f"⏭️ {symbol} пропущен — корреляция {corr_text} вне диапазона BUY")
                        return
                else:
                    if not apply_range_filter(corr_text, CORREL_MIN_SELL, CORREL_MAX_SELL):
                        print(f"⏭️ {symbol} пропущен — корреляция {corr_text} вне диапазона SELL")
                        return

            trade_id   = get_next_trade_id()
            strategies = {}
            for name, strat_cfg in STRATEGIES.items():
                if side == "BUY":
                    tp = entry_price * (1 + strat_cfg["tp"])
                    sl = entry_price * (1 - abs(strat_cfg["sl"]))
                else:
                    tp = entry_price * (1 - strat_cfg["tp"])
                    sl = entry_price * (1 + abs(strat_cfg["sl"]))
                strategies[name] = {"tp": tp, "sl": sl, "status": "OPEN"}

            with TRADES_LOCK:
                ACTIVE_TRADES[trade_id] = {
                    "symbol":      symbol,
                    "side":        side,
                    "entry_price": entry_price,
                    "strategies":  strategies,
                    "open_time":   datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                }
            save_active_trades()

            write_trade_to_excel(
                trade_id,
                {
                    "symbol":      symbol,
                    "signals":     res["signals"],
                    "strategies":  strategies,
                    "entry_price": entry_price,
                    "natr":        res["natr"],
                    "swing_num":   res["swing_num"],
                    "atr_ratio_50": res["atr_ratio_50"],
                    "atr_ratio_75": res["atr_ratio_75"],
                    "atr_ratio_100": res["atr_ratio_100"],
                    "expansion_3": res["expansion_3"],
                    "expansion_5": res["expansion_5"],
                },
                vol_text=res["volText"],
                vol24=res["volume_24h"] / 1_000_000,
                corr_text=corr_text
            )
            
            insert_trade(trade_id, BOT_NAME, {
                "symbol": symbol,
                "side": side,
                "entry_price": entry_price,
                "signals": res["signals"],
                "strategies": strategies,
                "natr": res["natr"],
                "swing_num": res["swing_num"],
                "delta_pct": res.get("delta_pct"),
                "relvol": res.get("relvol"),
                "atr_ratio_50": res["atr_ratio_50"],
                "atr_ratio_75": res["atr_ratio_75"],
                "atr_ratio_100": res["atr_ratio_100"],
                "expansion_3": res["expansion_3"],
                "expansion_5": res["expansion_5"],
            }, vol_text=res["volText"], vol_24h=res["volume_24h"]/1_000_000, corr_text=corr_text)

            vol24 = res["volume_24h"] / 1_000_000
            msg_text = (
                f"🤖 {BOT_NAME}\n"
                f"🔥 {res['symbol']}\n"
                f"Тип: {', '.join(res['signals'])}\n"
                f"Close: {res['close']:.6f}\n"
                f"EMA20: {res['ema20']:.6f}\n"
                f"EMA200: {res['ema200']:.6f}\n"
                f"Upper TL: {res['upper']}\n"
                f"Lower TL: {res['lower']}\n"
                f"VOL {res['volText']}\n"
                f"VOL 24h: {vol24:.1f}M USDT\n"
                f"Corr BTC: {corr_text}\n"
                f"NATR: {res['natr']}%\n"
                f"RelVol: {res['relvol']}\n"
                f"Свинг: {res['swing_num']}\n"
            )
            print(msg_text)
            # send_telegram для сигналов отключён

        except Exception as e:
            print(f"Ошибка process_signal: {e}")

    def handle_kline(msg):
        task_queue.put(msg)
        
    def handle_mark_price(msg):
        try:
            data = msg.get('data', msg)
            if data.get('e') != 'markPriceUpdate':
                return
            symbol = data.get('s')
            price  = float(data.get('p', 0))
            if price <= 0:
                return
            with TRADES_LOCK:
                has_open = any(
                    trade["symbol"] == symbol and
                    any(s["status"] == "OPEN" for s in trade["strategies"].values())
                    for trade in ACTIVE_TRADES.values()
                )
            if not has_open:
                return
            check_and_close_strategies(symbol, price, price)
        except Exception as e:
            print(f"Ошибка handle_mark_price: {e}")
    
    def worker():
        while True:
            msg = task_queue.get()
            process_signal(msg)
            task_queue.task_done()

    Thread(target=worker, daemon=True).start()

    chunk_size = 10

    while True:
        try:
            open_syms = get_symbols_with_open_trades()
            all_symbols = list(set(symbols) | open_syms)

            twm = ThreadedWebsocketManager()
            twm.start()

            for i in range(0, len(all_symbols), chunk_size):
                streams = [f"{s.lower()}@kline_1h" for s in all_symbols[i:i+chunk_size]]
                twm.start_multiplex_socket(callback=handle_kline, streams=streams)
            
            # markPrice для real-time TP/SL
            open_syms_now = get_symbols_with_open_trades()
            if open_syms_now:
                mark_twm = ThreadedWebsocketManager()
                mark_twm.start()
                for i in range(0, len(open_syms_now), chunk_size):
                    streams = [f"{s.lower()}@markPrice@1s" for s in list(open_syms_now)[i:i+chunk_size]]
                    mark_twm.start_multiplex_socket(callback=handle_mark_price, streams=streams)
                print(f"📡 markPrice запущен для {len(open_syms_now)} символов")
            
            print("🟢 WebSocket запущен")
            send_telegram(f"🟢 {BOT_NAME} WebSocket запущен")

            time.sleep(24 * 60 * 60)
            print("♻️ Плановый перезапуск WebSocket...")
            save_active_trades()
            twm.stop()

        except Exception as e:
            err_str = str(e)
            print(f"🔴 WebSocket упал: {err_str}. Переподключение через 30 секунд...")
            if "Socket Manager" not in err_str and "initialize" not in err_str:
                send_telegram(f"🔴 {BOT_NAME} WebSocket упал: {err_str}")
            save_active_trades()
            try:
                twm.stop()
            except Exception:
                pass
            time.sleep(30)


if __name__ == "__main__":
    main()
