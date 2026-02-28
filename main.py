from binance.client import Client
from binance import ThreadedWebsocketManager
import pandas as pd
import time
from datetime import datetime, timezone
import requests
import os
import json
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from threading import Thread, Lock
from queue import Queue

# ===== ЗАГРУЗКА КОНФИГА =====
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=True)
args = parser.parse_args()

with open(args.config, "r") as f:
    config = json.load(f)

BOT_NAME = config["NAME"]

load_dotenv()

# ================= НАСТРОЙКИ =================
MIN_24H_VOLUME = config["MIN_24H_VOLUME"]
LOOKBACK_CANDLES = config["LOOKBACK_CANDLES"]
VOLUME_LOOKBACK = config["VOLUME_LOOKBACK"]

VOL_MULT_TREND = float(config["VOL_MULT_TREND"])
VOL_MULT_COUNTER = float(config["VOL_MULT_COUNTER"])

EMA_FAST = config["EMA_FAST"]
EMA_SLOW = config["EMA_SLOW"]

MIN_BODY_TREND = float(config["MIN_BODY_TREND"])
MIN_BODY_COUNTER = float(config["MIN_BODY_COUNTER"])

ATR_LEN = config["ATR_LEN"]
ATR_GAP_MULT = float(config["ATR_GAP_MULT"])
EMA20_PROXIMITY_MULT = float(config["EMA20_PROXIMITY_MULT"])
EMA200_PROXIMITY_MULT = float(config["EMA200_PROXIMITY_MULT"])

COOLDOWN_BARS = config["COOLDOWN_BARS"]
USE_HTF_FILTER = config.get("USE_HTF_FILTER", False)  # фильтр старшего ТФ, по умолчанию выключен

BTC_LOOKBACK = config["BTC_LOOKBACK"]
EXCEL_STRAT_START_COL = 14  # колонка N в Excel
PREV_VOL_WINDOW = 3

CHAT_ID = os.getenv("CHAT_ID")
BOT_TOKEN = os.getenv("BOT_TOKEN")

client = Client()
BLACKLIST = {
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT"
}

# ================= TRADES =================
TRADE_STATE_FILE = f"trades_state_{BOT_NAME}.json"
EXCEL_FILE = f"trades_{BOT_NAME}.xlsx"

# FIX: блокировки для потокобезопасности
TRADES_LOCK = Lock()
EXCEL_LOCK = Lock()
_ID_LOCK = Lock()

SHEET_MAP = {
    "CONFIG_1": "config1",
    "CONFIG_2": "config2",
    "CONFIG_3": "config3",
    "CONFIG_4": "config4",
    "CONFIG_5": "config5",
    "CONFIG_6": "config6",
}

# FIX: 20:3 заменено на 4.5:1.5
STRATEGIES = {
    "3:1":    {"tp": 0.03,  "sl": -0.01},
    "6:1":    {"tp": 0.06,  "sl": -0.01},
    "6:2":    {"tp": 0.06,  "sl": -0.02},
    "10:3":   {"tp": 0.10,  "sl": -0.03},
    "4.5:1.5": {"tp": 0.045, "sl": -0.015},
}

def load_trade_id():
    if not os.path.exists(TRADE_STATE_FILE):
        return 0
    with open(TRADE_STATE_FILE, "r") as f:
        return json.load(f).get("last_trade_id", 0)

def save_trade_id(tid):
    with open(TRADE_STATE_FILE, "w") as f:
        json.dump({"last_trade_id": tid}, f)

# ================= ACTIVE TRADES PERSISTENCE =================
ACTIVE_TRADES_FILE = f"active_trades_{BOT_NAME}.json"

def save_active_trades():
    with TRADES_LOCK:
        with open(ACTIVE_TRADES_FILE, "w") as f:
            json.dump(ACTIVE_TRADES, f)

def load_active_trades():
    if not os.path.exists(ACTIVE_TRADES_FILE):
        return {}
    with open(ACTIVE_TRADES_FILE, "r") as f:
        return json.load(f)

# FIX: ACTIVE_TRADES загружается после определения функций
ACTIVE_TRADES = load_active_trades()

LAST_TRADE_ID = load_trade_id()

def get_next_trade_id():
    global LAST_TRADE_ID
    with _ID_LOCK:
        LAST_TRADE_ID += 1
        save_trade_id(LAST_TRADE_ID)
        return f"{LAST_TRADE_ID:05d}"

# ================= TELEGRAM =================
def send_telegram(message: str):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": message}
    try:
        requests.post(url, data=payload, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")

# ================= EXCEL =================
def write_trade_to_excel(trade_id, trade_info, vol_text, vol24, corr_text):
    sheet_name = SHEET_MAP.get(BOT_NAME, "config1")

    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            for sn in SHEET_MAP.values():
                if sn not in wb.sheetnames:
                    wb.create_sheet(sn)
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        headers = {
            "A":"Дата","B":"Время","C":"День","D":"Тикет","E":"Объем",
            "F":"Trade_id","G":"Тип","H":"Импульс","J":"Цена входа",
            "K":"Корреляция","M":"NATR%",
            "N":"3:1","O":"6:1","P":"6:2","Q":"10:3","R":"4.5:1.5",
            "S":"3:1 цена/PnL","T":"6:1 цена/PnL","U":"6:2 цена/PnL",
            "V":"10:3 цена/PnL","W":"4.5:1.5 цена/PnL"
        }
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            for col, header in headers.items():
                ws[f"{col}1"] = header

        next_row = ws.max_row + 1
        dt = datetime.now()
        ws["A"+str(next_row)] = dt.strftime("%d.%m.%Y")
        ws["B"+str(next_row)] = dt.strftime("%H:%M:%S")
        ws["C"+str(next_row)] = dt.strftime("%a")
        ws["D"+str(next_row)] = trade_info["symbol"]
        ws["E"+str(next_row)] = vol24
        ws["F"+str(next_row)] = trade_id
        ws["G"+str(next_row)] = ", ".join(trade_info["signals"])
        ws["H"+str(next_row)] = vol_text
        ws["J"+str(next_row)] = trade_info["entry_price"]
        ws["K"+str(next_row)] = corr_text
        ws["M"+str(next_row)] = trade_info["natr"]

        for idx, s in enumerate(STRATEGIES.keys()):
            col = get_column_letter(EXCEL_STRAT_START_COL + idx)
            ws[f"{col}{next_row}"] = trade_info["strategies"][s]["status"]

        wb.save(EXCEL_FILE)

def update_trade_status_in_excel(trade_id, strategy_name, status, close_price, pnl):
    sheet_name = SHEET_MAP.get(BOT_NAME, "config1")

    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]

        for row in range(2, ws.max_row+1):
            if str(ws[f"F{row}"].value) == trade_id:
                # статус в N-R
                col_map_status  = {"3:1":"N","6:1":"O","6:2":"P","10:3":"Q","4.5:1.5":"R"}
                # цена/PnL в S-W
                col_map_details = {"3:1":"S","6:1":"T","6:2":"U","10:3":"V","4.5:1.5":"W"}
                col_s = col_map_status[strategy_name]
                col_d = col_map_details[strategy_name]
                ws[f"{col_s}{row}"] = status
                ws[f"{col_d}{row}"] = f"{close_price:.6f} / {pnl:+.2f}%"
                break

        wb.save(EXCEL_FILE)

# ================= INDICATORS =================
def calculate_session_vwap(df):
    df = df.copy()
    df["date"] = pd.to_datetime(df["open_time"], unit="ms").dt.date
    tp = (df["high"] + df["low"] + df["close"])/3
    df["tpv"] = tp * df["volume"]
    df["cum_tpv"] = df.groupby("date")["tpv"].cumsum()
    df["cum_vol"] = df.groupby("date")["volume"].cumsum()
    return df["cum_tpv"]/df["cum_vol"]

def calculate_atr(df, period):
    hl = df["high"] - df["low"]
    hc = (df["high"] - df["close"].shift()).abs()
    lc = (df["low"] - df["close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def has_recent_spike(series, bars):
    if bars == 0:
        return False
    return series[-bars:].any()

def get_liquid_futures_symbols():
    tickers = client._request_futures_api(method="get", path="ticker/24hr")
    symbols = []
    for t in tickers:
        symbol = t["symbol"]
        if not symbol.endswith("USDT") or symbol in BLACKLIST:
            continue
        if float(t["quoteVolume"]) < MIN_24H_VOLUME:
            continue
        symbols.append(symbol)
    return symbols

def get_btc_returns():
    try:
        klines_btc = client.futures_klines(
            symbol="BTCUSDT", interval=Client.KLINE_INTERVAL_5MINUTE, limit=BTC_LOOKBACK
        )
        df_btc = pd.DataFrame(klines_btc, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_btc["close"] = df_btc["close"].astype(float)
        return df_btc["close"].pct_change()
    except Exception as e:
        print(f"Ошибка загрузки BTC свечей: {e}")
        return None

def check_volume_signal(symbol):
    klines = client.futures_klines(symbol=symbol,
        interval=Client.KLINE_INTERVAL_5MINUTE, limit=LOOKBACK_CANDLES)
    df = pd.DataFrame(klines, columns=[
        "open_time","open","high","low","close",
        "volume","close_time","quote_volume",
        "trades","taker_buy_base","taker_buy_quote","ignore"
    ])
    for c in ["open","high","low","close","volume"]:
        df[c] = df[c].astype(float)

    df["ema20"] = df["close"].ewm(span=EMA_FAST, adjust=False).mean()
    df["ema200"] = df["close"].ewm(span=EMA_SLOW, adjust=False).mean()
    df["atr"] = calculate_atr(df, ATR_LEN)
    df["natr"] = (df["atr"] / df["close"]) * 100
    df["vwap"] = calculate_session_vwap(df)
    df["quote_volume"] = df["close"]*df["volume"]
    avg_vol = df["quote_volume"].iloc[-(VOLUME_LOOKBACK+2):-2].mean()
    last = df.iloc[-2]

    spike_trend = last["quote_volume"] >= avg_vol*VOL_MULT_TREND
    spike_counter = last["quote_volume"] >= avg_vol*VOL_MULT_COUNTER

    body = abs(last["close"] - last["open"])
    rng = last["high"] - last["low"]
    body_pct = 0 if rng==0 else body/rng*100
    bull = last["close"]>last["open"]
    bear = last["close"]<last["open"]

    strong_body_trend = body_pct >= MIN_BODY_TREND
    strong_body_counter = body_pct >= MIN_BODY_COUNTER

    below_ema20 = last["open"]<last["ema20"] and last["close"]<last["ema20"]
    above_ema20 = last["open"]>last["ema20"] and last["close"]>last["ema20"]

    below_vwap = last["open"]<last["vwap"] and last["close"]<last["vwap"]
    above_vwap = last["open"]>last["vwap"] and last["close"]>last["vwap"]

    buy_low_condition = last["low"] < last["ema20"] and last["low"] < last["ema200"]
    sell_high_condition = last["high"] > last["ema20"] and last["high"] > last["ema200"]

    bull_trend = last["ema20"] > last["ema200"]
    bear_trend = last["ema20"] < last["ema200"]

    atr = last["atr"]
    emas_far_enough = abs(last["ema20"] - last["ema200"]) >= atr*ATR_GAP_MULT
    ema20_far_vwap = abs(last["ema20"] - last["vwap"]) >= atr*EMA20_PROXIMITY_MULT
    ema200_far_vwap = abs(last["ema200"] - last["vwap"]) >= atr*EMA200_PROXIMITY_MULT
    ema20_far_ema200 = abs(last["ema20"] - last["ema200"]) >= atr*EMA20_PROXIMITY_MULT
    ema20_clear_zone = ema20_far_vwap and ema20_far_ema200 and ema200_far_vwap

    # ================= HTF ФИЛЬТР (1ч) =================
    htf_bull = True
    htf_bear = True
    if USE_HTF_FILTER:
        try:
            klines_1h = client.futures_klines(symbol=symbol,
                interval=Client.KLINE_INTERVAL_1HOUR, limit=210)
            df_1h = pd.DataFrame(klines_1h, columns=[
                "open_time","open","high","low","close","volume",
                "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
            ])
            df_1h["close"] = df_1h["close"].astype(float)
            ema20_1h  = df_1h["close"].ewm(span=EMA_FAST, adjust=False).mean().iloc[-2]
            ema200_1h = df_1h["close"].ewm(span=EMA_SLOW, adjust=False).mean().iloc[-2]
            # Инвертированная логика — против тренда на 1ч
            htf_bull = ema20_1h < ema200_1h  # для BUY — на 1ч медвежий тренд
            htf_bear = ema20_1h > ema200_1h  # для SELL — на 1ч бычий тренд
        except Exception as e:
            print(f"Ошибка HTF фильтра {symbol}: {e}")

    signals = []
    if spike_trend and bull and strong_body_trend and below_ema20 and below_vwap and bull_trend and emas_far_enough and buy_low_condition and htf_bull:
        signals.append("BUY_TREND")
    if spike_trend and bear and strong_body_trend and above_ema20 and above_vwap and bear_trend and emas_far_enough and sell_high_condition and htf_bear:
        signals.append("SELL_TREND")
    if spike_counter and bull and strong_body_counter and below_ema20 and below_vwap and bear_trend and emas_far_enough and ema20_clear_zone and htf_bull:
        signals.append("BUY_COUNTER")
    if spike_counter and bear and strong_body_counter and above_ema20 and above_vwap and bull_trend and emas_far_enough and ema20_clear_zone and htf_bear:
        signals.append("SELL_COUNTER")

    if not signals:
        return None

    ticker_24h = client.futures_ticker(symbol=symbol)
    volume_24h = float(ticker_24h["quoteVolume"])

    return {
        "symbol": symbol,
        "signals": signals,
        "close": last["close"],
        "low": last["low"],
        "high": last["high"],
        "ema20": last["ema20"],
        "ema200": last["ema200"],
        "vwap": last["vwap"],
        "natr": round(last["natr"], 3),
        "volText": f"x{last['quote_volume']/avg_vol:.2f}",
        "prevVolCount": int((df.iloc[-5:-2]["quote_volume"] > last["quote_volume"]).sum()),
        "volume_24h": volume_24h
    }

# ================= MAIN =================
def main():
    symbols = get_liquid_futures_symbols()
    print(f"✅ Ликвидные токены: {len(symbols)}")

    last_signal_time = {}
    cooldown_seconds = COOLDOWN_BARS * 5 * 60

    def update_symbols_periodically():
        nonlocal symbols
        while True:
            time.sleep(3600)
            try:
                symbols = get_liquid_futures_symbols()
                print(f"♻️ Обновление токенов: {len(symbols)}")
            except Exception as e:
                print(f"Ошибка обновления токенов: {e}")

    Thread(target=update_symbols_periodically, daemon=True).start()

    task_queue = Queue()

    def process_signal(msg):
        try:
            if msg.get("e") == "error":
                print(f"🔴 WebSocket ошибка: {msg}")
                send_telegram(f"🔴 {BOT_NAME} WebSocket ошибка: {msg.get('m', 'неизвестно')}")
                return

            if 'data' not in msg or 'k' not in msg['data']:
                return
            candle = msg['data']['k']
            symbol = candle['s']
            if symbol not in symbols or not candle['x']:
                return

            price_high = float(candle["h"])
            price_low = float(candle["l"])

            # ===== Закрытие открытых стратегий =====
            closed_trades = []
            with TRADES_LOCK:
                for trade_id, trade in list(ACTIVE_TRADES.items()):
                    if trade["symbol"] != symbol:
                        continue
                    for strat_name, strat in trade["strategies"].items():
                        if strat["status"] != "OPEN":
                            continue
                        if trade["side"] == "BUY":
                            if price_low <= strat["sl"]:
                                result = "SL"
                            elif price_high >= strat["tp"]:
                                result = "TP"
                            else:
                                continue
                        else:
                            if price_high >= strat["sl"]:
                                result = "SL"
                            elif price_low <= strat["tp"]:
                                result = "TP"
                            else:
                                continue

                        strat["status"] = result
                        # Цена закрытия и PnL
                        close_price = strat["sl"] if result == "SL" else strat["tp"]
                        pnl = (close_price - trade["entry_price"]) / trade["entry_price"] * 100
                        if trade["side"] == "SELL":
                            pnl = -pnl
                        pnl = round(pnl, 2)
                        msg_text = f"📊 TRADE CLOSED\nID:{trade_id}\n{trade['symbol']} {trade['side']}\nStrategy:{strat_name}\nEntry:{trade['entry_price']:.6f}\nClose:{close_price:.6f}\nPnL:{pnl:+.2f}%\n{result}"
                        Thread(target=update_trade_status_in_excel, args=(trade_id, strat_name, result, close_price, pnl), daemon=True).start()
                        Thread(target=send_telegram, args=(msg_text,), daemon=True).start()
                        print(msg_text)

                    if all(s["status"] != "OPEN" for s in trade["strategies"].values()):
                        closed_trades.append(trade_id)

                for tid in closed_trades:
                    del ACTIVE_TRADES[tid]

            # FIX: сохраняем после удаления закрытых трейдов
            if closed_trades:
                save_active_trades()

            # Cooldown
            now = time.time()
            if now - last_signal_time.get(symbol, 0) < cooldown_seconds:
                return

            # ===== Новые сигналы =====
            res = check_volume_signal(symbol)
            if not res:
                return

            last_signal_time[symbol] = now

            entry_price = res["close"]
            side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"

            # ===== Корреляция BTC =====
            try:
                btc_returns = get_btc_returns()
                if btc_returns is not None:
                    klines_sym = client.futures_klines(symbol=symbol, interval=Client.KLINE_INTERVAL_5MINUTE, limit=BTC_LOOKBACK)
                    df_sym = pd.DataFrame(klines_sym, columns=[
                        "open_time","open","high","low","close","volume",
                        "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
                    ])
                    df_sym["close"] = df_sym["close"].astype(float)
                    symbol_returns = df_sym["close"].pct_change()
                    btc_subset = btc_returns[-len(symbol_returns):]
                    corr = btc_subset.corr(symbol_returns)
                    corr_text = f"{corr:.2f}" if corr is not None else "N/A"
                else:
                    corr_text = "N/A"
            except Exception as e:
                print(f"Ошибка корреляции {symbol}: {e}")
                corr_text = "N/A"

            trade_id = get_next_trade_id()
            strategies = {}
            for name, strat_cfg in STRATEGIES.items():
                if side == "BUY":
                    tp = entry_price * (1 + strat_cfg["tp"])
                    sl = entry_price * (1 - abs(strat_cfg["sl"]))
                else:
                    tp = entry_price * (1 - strat_cfg["tp"])
                    sl = entry_price * (1 + abs(strat_cfg["sl"]))
                strategies[name] = {"tp": tp, "sl": sl, "status": "OPEN"}

            # FIX: потокобезопасное добавление + сохранение
            with TRADES_LOCK:
                ACTIVE_TRADES[trade_id] = {
                    "symbol": symbol,
                    "side": side,
                    "entry_price": entry_price,
                    "strategies": strategies,
                    "open_time": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                }
            save_active_trades()

            write_trade_to_excel(
                trade_id,
                {
                    "symbol": symbol,
                    "signals": res["signals"],
                    "strategies": strategies,
                    "entry_price": entry_price,
                    "natr": res["natr"]
                },
                vol_text=res["volText"],
                vol24=res["volume_24h"]/1_000_000,
                corr_text=corr_text
            )

            # ===== Telegram =====
            vol24 = res["volume_24h"]/1_000_000
            msg_text = (
                f"🤖 {BOT_NAME}\n"
                f"🔥 {res['symbol']}\n"
                f"Тип: {', '.join(res['signals'])}\n"
                f"Close: {res['close']:.6f}\n"
                f"EMA20: {res['ema20']:.6f}\n"
                f"EMA200: {res['ema200']:.6f}\n"
                f"VWAP: {res['vwap']:.6f}\n"
                f"VOL {res['volText']}\n"
                f"Prev volume higher: {res['prevVolCount']}/3\n"
                f"VOL 24h: {vol24:.1f}M USDT\n"
                f"Corr BTC: {corr_text}\n"
                f"NATR: {res['natr']}%\n"
            )
            print(msg_text)
            send_telegram(msg_text)

        except Exception as e:
            print(f"Ошибка process_signal: {e}")

    def handle_kline(msg):
        task_queue.put(msg)

    def worker():
        while True:
            msg = task_queue.get()
            process_signal(msg)
            task_queue.task_done()

    Thread(target=worker, daemon=True).start()

    # ===== WebSocket с переподключением и плановым перезапуском =====
    chunk_size = 30

    while True:
        try:
            twm = ThreadedWebsocketManager()
            twm.start()

            for i in range(0, len(symbols), chunk_size):
                streams = [f"{s.lower()}@kline_5m" for s in symbols[i:i+chunk_size]]
                twm.start_multiplex_socket(callback=handle_kline, streams=streams)

            print("🟢 WebSocket запущен")
            send_telegram(f"🟢 {BOT_NAME} WebSocket запущен")

            # Плановый перезапуск каждые 24 часа
            time.sleep(24 * 60 * 60)
            print("♻️ Плановый перезапуск WebSocket...")
            send_telegram(f"♻️ {BOT_NAME} плановый перезапуск WebSocket")
            save_active_trades()
            twm.stop()

        except Exception as e:
            print(f"🔴 WebSocket упал: {e}. Переподключение через 30 секунд...")
            send_telegram(f"🔴 {BOT_NAME} WebSocket упал: {e}. Переподключение через 30 секунд...")
            save_active_trades()
            try:
                twm.stop()
            except Exception:
                pass
            time.sleep(30)


if __name__ == "__main__":
    main()
