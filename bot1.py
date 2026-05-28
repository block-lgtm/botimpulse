from binance.client import Client
from binance import ThreadedWebsocketManager
import pandas as pd
import time
from datetime import datetime, UTC
import requests
import os
import json
import argparse
import openpyxl
import math
import asyncio
from db import init_db, insert_trade, update_strategy_status
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from threading import Thread, Lock
from queue import Queue

# ===== ЗАГРУЗКА КОНФИГА =====
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=True)
args = parser.parse_args()

with open(args.config, "r") as f:
    config = json.load(f)

BOT_NAME = config["NAME"]

load_dotenv()

# Фильтр по объёму свечи (quote volume текущей свечи, не 24h)
USE_VOLUME_FILTER  = config.get("USE_VOLUME_FILTER", False)
VOLUME_MIN_BUY     = float(config.get("VOLUME_MIN_BUY",  0))
VOLUME_MAX_BUY     = float(config.get("VOLUME_MAX_BUY",  0))
VOLUME_MIN_SELL    = float(config.get("VOLUME_MIN_SELL", 0))
VOLUME_MAX_SELL    = float(config.get("VOLUME_MAX_SELL", 0))

# Фильтр по NATR
USE_NATR_FILTER    = config.get("USE_NATR_FILTER", False)
NATR_MIN_BUY       = float(config.get("NATR_MIN_BUY",  0))
NATR_MAX_BUY       = float(config.get("NATR_MAX_BUY",  0))
NATR_MIN_SELL      = float(config.get("NATR_MIN_SELL", 0))
NATR_MAX_SELL      = float(config.get("NATR_MAX_SELL", 0))

# Фильтр по корреляции с BTC
USE_CORREL_FILTER  = config.get("USE_CORREL_FILTER", False)
CORREL_MIN_BUY     = float(config.get("CORREL_MIN_BUY",  0))
CORREL_MAX_BUY     = float(config.get("CORREL_MAX_BUY",  0))
CORREL_MIN_SELL    = float(config.get("CORREL_MIN_SELL", 0))
CORREL_MAX_SELL    = float(config.get("CORREL_MAX_SELL", 0))

# Дни недели когда не торгуем
SKIP_DAYS          = [d.lower() for d in config.get("SKIP_DAYS", [])]
# Пример: ["monday", "saturday", "sunday"]

# ================= НАСТРОЙКИ =================
MIN_24H_VOLUME   = config["MIN_24H_VOLUME"]
LOOKBACK_CANDLES = config["LOOKBACK_CANDLES"]
VOLUME_LOOKBACK  = config["VOLUME_LOOKBACK"]

VOL_MULT         = float(config["VOL_MULT"])
MIN_BODY_PCT     = float(config["MIN_BODY_PCT"])
COOLDOWN_BARS    = config["COOLDOWN_BARS"]

EMA_FAST         = config["EMA_FAST"]
EMA_SLOW         = config["EMA_SLOW"]

BTC_LOOKBACK     = config["BTC_LOOKBACK"]
ATR_LEN          = config["ATR_LEN"]

# Фильтры — включить/выключить
USE_EMA_FILTER  = config.get("USE_EMA_FILTER", True)
USE_VWAP_FILTER = config.get("USE_VWAP_FILTER", True)

# Свинг фильтры (0 = выключен)
SWING_BUY_TREND    = config.get("SWING_BUY_TREND", 0)
SWING_SELL_TREND   = config.get("SWING_SELL_TREND", 0)
SWING_BUY_COUNTER  = config.get("SWING_BUY_COUNTER", 0)
SWING_SELL_COUNTER = config.get("SWING_SELL_COUNTER", 0)

# Delta фильтр
USE_DELTA_FILTER   = config.get("USE_DELTA_FILTER", False)
DELTA_MIN_BUY      = float(config.get("DELTA_MIN_BUY", 0.0))   # 0 = просто положительная
DELTA_MAX_BUY      = float(config.get("DELTA_MAX_BUY", 0.0))   # 0 = без верхнего лимита
DELTA_MIN_SELL     = float(config.get("DELTA_MIN_SELL", 0.0))  # 0 = просто отрицательная
DELTA_MAX_SELL     = float(config.get("DELTA_MAX_SELL", 0.0))  # 0 = без нижнего лимита

# RelVol фильтр
USE_RELVOL_FILTER  = config.get("USE_RELVOL_FILTER", False)
RELVOL_ANCHOR      = config.get("RELVOL_ANCHOR", "1d")   # "1h", "4h", "1d", "1w"
RELVOL_LENGTH      = config.get("RELVOL_LENGTH", 5)       # кол-во периодов anchor
RELVOL_MIN_BUY     = float(config.get("RELVOL_MIN_BUY",  0.0))  # 0 = выкл
RELVOL_MAX_BUY     = float(config.get("RELVOL_MAX_BUY",  0.0))  # 0 = выкл
RELVOL_MIN_SELL    = float(config.get("RELVOL_MIN_SELL", 0.0))  # 0 = выкл
RELVOL_MAX_SELL    = float(config.get("RELVOL_MAX_SELL", 0.0))  # 0 = выкл

ANCHOR_MAP = {
    "1h": Client.KLINE_INTERVAL_1HOUR,
    "4h": Client.KLINE_INTERVAL_4HOUR,
    "1d": Client.KLINE_INTERVAL_1DAY,
    "1w": Client.KLINE_INTERVAL_1WEEK,
}

EXCEL_STRAT_START_COL = 14  # колонка N
PREV_VOL_WINDOW  = 3

CHAT_ID   = os.getenv("CHAT_ID")
BOT_TOKEN = os.getenv("BOT_TOKEN")

client = Client()
BLACKLIST = {
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT"
}

# ================= TRADES =================
TRADE_STATE_FILE   = f"trades_state_{BOT_NAME}.json"
EXCEL_FILE         = f"trades_{BOT_NAME}.xlsx"
ACTIVE_TRADES_FILE = f"active_trades_{BOT_NAME}.json"

TRADES_LOCK = Lock()
EXCEL_LOCK  = Lock()
_ID_LOCK    = Lock()

SHEET_MAP = {
    "CONFSP1":  "confsp1",
}

# Стратегии: 3:1, 6:1, 6:2, 10:3, 12:4
_strat_raw = config.get("STRATEGIES_CONFIG", {
    "3:1":  {"enabled": True, "tp": 0.03, "sl": 0.01, "BUY": {}, "SELL": {}},
    "6:1":  {"enabled": True, "tp": 0.06, "sl": 0.01, "BUY": {}, "SELL": {}},
    "6:2":  {"enabled": True, "tp": 0.06, "sl": 0.02, "BUY": {}, "SELL": {}},
    "10:3": {"enabled": True, "tp": 0.10, "sl": 0.03, "BUY": {}, "SELL": {}},
    "12:4": {"enabled": True, "tp": 0.12, "sl": 0.04, "BUY": {}, "SELL": {}},
})

STRATEGIES = {
    name: {
        "tp":   float(s["tp"]),
        "sl":   -float(s["sl"]),
        "BUY":  s.get("BUY",  {}),
        "SELL": s.get("SELL", {}),
    }
    for name, s in _strat_raw.items()
    if s.get("enabled", True)
}

print(f"✅ Активные стратегии: {list(STRATEGIES.keys())}")

def check_and_close_strategies(symbol, price_high, price_low):
    closed_trades = []
    with TRADES_LOCK:
        for trade_id, trade in list(ACTIVE_TRADES.items()):
            if trade["symbol"] != symbol:
                continue
            for strat_name, strat in trade["strategies"].items():
                if strat["status"] != "OPEN":
                    continue
                result = None
                if trade["side"] == "BUY":
                    if price_low <= strat["sl"]:
                        result = "SL"
                    elif price_high >= strat["tp"]:
                        result = "TP"
                else:
                    if price_high >= strat["sl"]:
                        result = "SL"
                    elif price_low <= strat["tp"]:
                        result = "TP"
                if result:
                    strat["status"] = result
                    strat["close_time"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
                    update_strategy_status(trade_id, strat_name, result)
                    Thread(
                        target=update_trade_status_in_excel,
                        args=(trade_id, strat_name, result),
                        daemon=True
                    ).start()
            if all(s["status"] != "OPEN" for s in trade["strategies"].values()):
                closed_trades.append(trade_id)
        for tid in closed_trades:
            del ACTIVE_TRADES[tid]
    if closed_trades:
        save_active_trades()
    return closed_trades

def load_trade_id():
    if not os.path.exists(TRADE_STATE_FILE):
        return 0
    with open(TRADE_STATE_FILE, "r") as f:
        return json.load(f).get("last_trade_id", 0)

def save_trade_id(tid):
    with open(TRADE_STATE_FILE, "w") as f:
        json.dump({"last_trade_id": tid}, f)

def save_active_trades():
    with TRADES_LOCK:
        with open(ACTIVE_TRADES_FILE, "w") as f:
            json.dump(ACTIVE_TRADES, f)

def load_active_trades():
    if not os.path.exists(ACTIVE_TRADES_FILE):
        return {}
    with open(ACTIVE_TRADES_FILE, "r") as f:
        return json.load(f)

ACTIVE_TRADES = load_active_trades()
LAST_TRADE_ID = load_trade_id()

def get_next_trade_id():
    global LAST_TRADE_ID
    with _ID_LOCK:
        LAST_TRADE_ID += 1
        save_trade_id(LAST_TRADE_ID)
        # Префикс SP для volumespike, TL для trendline
        prefix = "SP" if "SP" in BOT_NAME else "TL"
        return f"{prefix}{LAST_TRADE_ID:05d}"

# ================= TELEGRAM =================
def send_telegram(message: str):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": message}
    try:
        requests.post(url, data=payload, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")

# ================= EXCEL =================
def write_trade_to_excel(trade_id, trade_info, vol_text, vol24, corr_text):
    sheet_name = SHEET_MAP.get(BOT_NAME, "confsp1")

    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            for sn in SHEET_MAP.values():
                if sn not in wb.sheetnames:
                    wb.create_sheet(sn)
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.active = wb[sheet_name]
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        headers = {
            "I":"Дата","B":"Время","C":"День","D":"Тикет","E":"Объем",
            "F":"Trade_id","G":"Тип","H":"Импульс","J":"Цена входа",
            "L":"Корреляция","M":"NATR%",
            "N":"3:1","O":"6:1","P":"6:2","Q":"10:3","R":"12:4",
            "S":"ATR Ratio 50","T":"ATR Ratio 75","U":"ATR Ratio 100",
            "V":"Expansion 3","W":"Expansion 5",
            "X":"Свинг","Y":"Delta%","Z":"RelVol"
        }
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            for col, header in headers.items():
                ws[f"{col}1"] = header

        next_row = ws.max_row + 1
        dt = datetime.now()
        ws["I"+str(next_row)] = dt.strftime("%d.%m.%Y")
        ws["B"+str(next_row)] = dt.strftime("%H:%M:%S")
        ws["C"+str(next_row)] = dt.strftime("%a")
        ws["D"+str(next_row)] = trade_info["symbol"]
        ws["E"+str(next_row)] = vol24
        ws["F"+str(next_row)] = trade_id
        ws["G"+str(next_row)] = ", ".join(trade_info["signals"])
        ws["H"+str(next_row)] = vol_text
        ws["J"+str(next_row)] = trade_info["entry_price"]
        ws["L"+str(next_row)] = corr_text
        ws["M"+str(next_row)] = trade_info["natr"]
        ws["S"+str(next_row)] = trade_info["atr_ratio_50"]
        ws["T"+str(next_row)] = trade_info["atr_ratio_75"]
        ws["U"+str(next_row)] = trade_info["atr_ratio_100"]
        ws["V"+str(next_row)] = trade_info["expansion_3"]
        ws["W"+str(next_row)] = trade_info["expansion_5"]
        ws["X"+str(next_row)] = trade_info["swing_num"]
        ws["Y"+str(next_row)] = trade_info["delta_pct"]
        ws["Z"+str(next_row)] = trade_info["relvol"]

        for idx, s in enumerate(STRATEGIES.keys()):
            col = get_column_letter(EXCEL_STRAT_START_COL + idx)
            ws[f"{col}{next_row}"] = trade_info["strategies"][s]["status"]
        wb.active = wb[sheet_name]
        wb.save(EXCEL_FILE)

def update_trade_status_in_excel(trade_id, strategy_name, status):
    sheet_name = SHEET_MAP.get(BOT_NAME, "confsp1")

    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]

        for row in range(2, ws.max_row+1):
            if str(ws[f"F{row}"].value) == trade_id:
                col_map_status  = {"3:1":"N","6:1":"O","6:2":"P","10:3":"Q","12:4":"R"}
                col_s = col_map_status[strategy_name]
                ws[f"{col_s}{row}"] = status
                break
        wb.active = wb[sheet_name]
        wb.save(EXCEL_FILE)

# ================= INDICATORS =================
def calculate_session_vwap(df):
    df = df.copy()
    df["date"] = pd.to_datetime(df["open_time"], unit="ms").dt.date
    tp = (df["high"] + df["low"] + df["close"]) / 3
    df["tpv"] = tp * df["volume"]
    df["cum_tpv"] = df.groupby("date")["tpv"].cumsum()
    df["cum_vol"] = df.groupby("date")["volume"].cumsum()
    return df["cum_tpv"] / df["cum_vol"]

def calculate_atr(df, period):
    hl = df["high"] - df["low"]
    hc = (df["high"] - df["close"].shift()).abs()
    lc = (df["low"] - df["close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def calculate_atr_ratio(df, avg_len, atr_len=14):
    atr = calculate_atr(df, atr_len)
    atr_avg = atr.rolling(avg_len).mean()

    if pd.isna(atr_avg.iloc[-2]) or atr_avg.iloc[-2] == 0:
        return 0

    ratio = atr.iloc[-2] / atr_avg.iloc[-2]
    return round(ratio, 4)


def calculate_expansion(df, lookback, natr_len=14):
    atr = calculate_atr(df, natr_len)

    last_close = df["close"].iloc[-2]
    prev_close = df["close"].iloc[-2 - lookback]

    atr_value = atr.iloc[-2]

    if prev_close == 0 or last_close == 0 or atr_value == 0:
        return 0

    # Pine:
    # natr = atr / close * 100
    natr = (atr_value / last_close) * 100

    # Pine:
    # abs(close - close[lookback]) / close[lookback] * 100
    price_change_pct = abs(last_close - prev_close) / prev_close * 100

    # Pine:
    # expansion = priceChangePct / (natr * sqrt(lookback))
    expansion = price_change_pct / (natr * math.sqrt(lookback))

    return round(expansion, 4)

def get_liquid_futures_symbols():
    tickers = client._request_futures_api(method="get", path="ticker/24hr")
    symbols = []
    for t in tickers:
        symbol = t["symbol"]
        if not symbol.endswith("USDT") or symbol in BLACKLIST:
            continue
        if float(t["quoteVolume"]) < MIN_24H_VOLUME:
            continue
        symbols.append(symbol)
    return symbols

def get_symbols_with_open_trades():
    with TRADES_LOCK:
        return {
            trade["symbol"] for trade in ACTIVE_TRADES.values()
            if any(s["status"] == "OPEN" for s in trade["strategies"].values())
        }

def get_btc_returns():
    try:
        klines_btc = client.futures_klines(
            symbol="BTCUSDT", interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
        )
        df_btc = pd.DataFrame(klines_btc, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_btc["close"] = df_btc["close"].astype(float)
        return df_btc["close"].pct_change()
    except Exception as e:
        print(f"Ошибка загрузки BTC свечей: {e}")
        return None

def check_swing(df, side, n):
    """
    Фильтр: True если сигнал НЕ должен быть срезан.
    BUY: ни одна из n предыдущих свечей не имеет low ниже текущей
    SELL: ни одна из n предыдущих свечей не имеет high выше текущей
    """
    if n == 0:
        return True
    current = df.iloc[-2]
    for i in range(1, n + 1):
        idx = -2 - i
        if abs(idx) > len(df):
            break
        candle = df.iloc[idx]
        if side == "BUY" and candle["low"] < current["low"]:
            return False
        if side == "SELL" and candle["high"] > current["high"]:
            return False
    return True

def get_swing_num(df, side, n=5):
    """
    Информационно: наидальнейшая свеча из n предыдущих
    у которой low ниже (BUY) или high выше (SELL) текущей.
    Возвращает номер дальней или 0.
    """
    current = df.iloc[-2]
    result = 0
    for i in range(1, n + 1):
        idx = -2 - i
        if abs(idx) > len(df):
            break
        candle = df.iloc[idx]
        if side == "BUY" and candle["low"] < current["low"]:
            result = i  # перезаписываем — остаётся дальняя
        if side == "SELL" and candle["high"] > current["high"]:
            result = i  # перезаписываем — остаётся дальняя
    return result

def calculate_delta(last_row):
    """Delta по формуле: buyVol = volume * (close-low)/(high-low)"""
    rng = last_row["high"] - last_row["low"]
    if rng == 0:
        return 0.0
    buy_vol  = last_row["volume"] * (last_row["close"] - last_row["low"]) / rng
    sell_vol = last_row["volume"] - buy_vol
    return round((buy_vol - sell_vol) / last_row["volume"] * 100, 2)

def calculate_relvol(symbol, current_vol, anchor_interval, length):
    """
    RelVol: текущий объём / средний объём за тот же час за последние N периодов anchor.
    """
    try:
        # Загружаем anchor свечи — нужно length+1 чтобы найти тот же час
        anchor_limit = length * 24 + 5 if anchor_interval == Client.KLINE_INTERVAL_1HOUR else length + 5
        klines = client.futures_klines(
            symbol=symbol, interval=anchor_interval, limit=anchor_limit
        )
        df_anchor = pd.DataFrame(klines, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_anchor["open_time"] = pd.to_datetime(df_anchor["open_time"], unit="ms")
        df_anchor["quote_volume"] = df_anchor["close"].astype(float) * df_anchor["volume"].astype(float)

        # Текущий час
        current_hour = df_anchor["open_time"].iloc[-2].hour

        if anchor_interval == Client.KLINE_INTERVAL_1HOUR:
            # Ищем свечи с тем же часом за последние N дней
            same_hour = df_anchor[df_anchor["open_time"].dt.hour == current_hour].iloc[-(length+1):-1]
        else:
            # Для дневных/недельных — просто последние N свечей кроме текущей
            same_hour = df_anchor.iloc[-(length+1):-1]

        if len(same_hour) == 0:
            return None

        avg_vol = same_hour["quote_volume"].mean()
        if avg_vol == 0:
            return None

        return round(current_vol / avg_vol, 3)
    except Exception as e:
        print(f"Ошибка RelVol {symbol}: {e}")
        return None

def apply_range_filter(value, min_val, max_val):
    if value is None:
        return False
    # min_val: 0 = выключен, иначе применяем
    if min_val != 0 and value < min_val:
        return False
    # max_val: 0 = выключен, иначе применяем
    if max_val != 0 and value > max_val:
        return False
    return True

def check_strategy_filters(strat_cfg, side, natr, delta_pct, candle_vol_m, corr):
    """
    Проверяет все фильтры конкретной стратегии для конкретной стороны (BUY/SELL).
    Возвращает True если сделка проходит все фильтры.
    """
    f = strat_cfg.get(side, {})
    if not f:
        return True  # нет фильтров — пропускаем

    today = datetime.now(UTC).strftime("%A").lower()
    skip_days = [d.lower() for d in f.get("SKIP_DAYS", [])]
    if today in skip_days:
        return False

    if f.get("USE_NATR_FILTER", False):
        if not apply_range_filter(natr, f.get("NATR_MIN", 0), f.get("NATR_MAX", 0)):
            return False

    if f.get("USE_DELTA_FILTER", False):
        if side == "BUY":
            if not apply_range_filter(delta_pct, f.get("DELTA_MIN", 0), f.get("DELTA_MAX", 0)):
                return False
        else:
            # SELL: delta должна быть отрицательной
            if not apply_range_filter(-delta_pct, f.get("DELTA_MIN", 0), f.get("DELTA_MAX", 0)):
                return False

    if f.get("USE_VOLUME_FILTER", False):
        if not apply_range_filter(candle_vol_m, f.get("VOLUME_MIN", 0), f.get("VOLUME_MAX", 0)):
            return False

    if f.get("USE_CORREL_FILTER", False) and corr is not None:
        try:
            corr_val = float(corr)
            if not apply_range_filter(corr_val, f.get("CORREL_MIN", 0), f.get("CORREL_MAX", 0)):
                return False
        except (ValueError, TypeError):
            pass

    return True

def check_volume_signal(symbol):
    klines = client.futures_klines(
        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=LOOKBACK_CANDLES
    )
    df = pd.DataFrame(klines, columns=[
        "open_time","open","high","low","close",
        "volume","close_time","quote_volume",
        "trades","taker_buy_base","taker_buy_quote","ignore"
    ])
    for c in ["open","high","low","close","volume"]:
        df[c] = df[c].astype(float)

    df["ema20"]  = df["close"].ewm(span=EMA_FAST, adjust=False).mean()
    df["ema200"] = df["close"].ewm(span=EMA_SLOW, adjust=False).mean()
    df["atr"]    = calculate_atr(df, ATR_LEN)
    df["natr"]   = (df["atr"] / df["close"]) * 100
    df["vwap"]   = calculate_session_vwap(df)
    df["quote_volume"] = df["close"] * df["volume"]

    atr_ratio_50  = calculate_atr_ratio(df, 50)
    atr_ratio_75  = calculate_atr_ratio(df, 75)
    atr_ratio_100 = calculate_atr_ratio(df, 100)

    expansion_3 = calculate_expansion(df, 3)
    expansion_5 = calculate_expansion(df, 5)

    avg_vol = df["quote_volume"].iloc[-(VOLUME_LOOKBACK + 2):-2].mean()
    last = df.iloc[-2]

    volume_spike = last["quote_volume"] >= avg_vol * VOL_MULT

    body     = abs(last["close"] - last["open"])
    rng      = last["high"] - last["low"]
    body_pct = 0 if rng == 0 else body / rng * 100
    bull = last["close"] > last["open"]
    bear = last["close"] < last["open"]

    strong_body = body_pct >= MIN_BODY_PCT

    # EMA фильтр
    bull_trend = last["ema20"] > last["ema200"]
    bear_trend = last["ema20"] < last["ema200"]
    ema_bull_ok = bull_trend if USE_EMA_FILTER else True
    ema_bear_ok = bear_trend if USE_EMA_FILTER else True

    # VWAP фильтр
    below_vwap = (last["close"] < last["vwap"]) if USE_VWAP_FILTER else True
    above_vwap = (last["close"] > last["vwap"]) if USE_VWAP_FILTER else True

    # Cooldown
    if COOLDOWN_BARS > 0:
        recent = df.iloc[-(COOLDOWN_BARS + 2):-2]
        recent_spike = (recent["quote_volume"] >= avg_vol * VOL_MULT).any()
    else:
        recent_spike = False

    # Свинг фильтры
    swing_buy_trend_ok    = check_swing(df, "BUY",  SWING_BUY_TREND)
    swing_sell_trend_ok   = check_swing(df, "SELL", SWING_SELL_TREND)

    # ===== Delta =====
    delta_pct = calculate_delta(last)
    if USE_DELTA_FILTER:
        # BUY: delta >= min (0 = просто положительная), <= max (0 = без лимита)
        delta_buy_ok  = (delta_pct > 0 if DELTA_MIN_BUY == 0 else delta_pct >= DELTA_MIN_BUY)
        if DELTA_MAX_BUY > 0:
            delta_buy_ok = delta_buy_ok and delta_pct <= DELTA_MAX_BUY
        # SELL: delta <= -min (0 = просто отрицательная), >= -max (0 = без лимита)
        delta_sell_ok = (delta_pct < 0 if DELTA_MIN_SELL == 0 else delta_pct <= -DELTA_MIN_SELL)
        if DELTA_MAX_SELL > 0:
            delta_sell_ok = delta_sell_ok and delta_pct >= -DELTA_MAX_SELL
    else:
        delta_buy_ok  = True
        delta_sell_ok = True

    # ===== RelVol =====
    anchor_interval = ANCHOR_MAP.get(RELVOL_ANCHOR, Client.KLINE_INTERVAL_1DAY)
    relvol = calculate_relvol(symbol, last["quote_volume"], anchor_interval, RELVOL_LENGTH) if USE_RELVOL_FILTER else None
    if USE_RELVOL_FILTER and relvol is not None:
        relvol_buy_ok  = (RELVOL_MIN_BUY  <= 0 or relvol >= RELVOL_MIN_BUY)  and (RELVOL_MAX_BUY  <= 0 or relvol <= RELVOL_MAX_BUY)
        relvol_sell_ok = (RELVOL_MIN_SELL <= 0 or relvol >= RELVOL_MIN_SELL) and (RELVOL_MAX_SELL <= 0 or relvol <= RELVOL_MAX_SELL)
    else:
        relvol_buy_ok  = True
        relvol_sell_ok = True

    # ===== Volume фильтр (quote volume текущей свечи в млн) =====
    candle_vol_m = last["quote_volume"] / 1_000_000
    if USE_VOLUME_FILTER:
        vol_buy_ok  = apply_range_filter(candle_vol_m, VOLUME_MIN_BUY,  VOLUME_MAX_BUY)
        vol_sell_ok = apply_range_filter(candle_vol_m, VOLUME_MIN_SELL, VOLUME_MAX_SELL)
    else:
        vol_buy_ok  = True
        vol_sell_ok = True

    # ===== NATR фильтр =====
    natr_val = last["natr"]
    if USE_NATR_FILTER:
        natr_buy_ok  = apply_range_filter(natr_val, NATR_MIN_BUY,  NATR_MAX_BUY)
        natr_sell_ok = apply_range_filter(natr_val, NATR_MIN_SELL, NATR_MAX_SELL)
    else:
        natr_buy_ok  = True
        natr_sell_ok = True

    signals = []
    if volume_spike and bull and strong_body and ema_bull_ok and below_vwap and not recent_spike and swing_buy_trend_ok and delta_buy_ok and relvol_buy_ok and vol_buy_ok and natr_buy_ok:
        signals.append("BUY_TREND")

    if volume_spike and bear and strong_body and ema_bear_ok and above_vwap and not recent_spike and swing_sell_trend_ok and delta_sell_ok and relvol_sell_ok and vol_sell_ok and natr_sell_ok:
        signals.append("SELL_TREND")

    if not signals:
        return None

    # Колонка X — наидальнейшая свеча из 5 (дальняя, не ближняя)
    side_for_swing = "BUY" if any("BUY" in s for s in signals) else "SELL"
    swing_num = get_swing_num(df, side_for_swing, 5)

    ticker_24h = client.futures_ticker(symbol=symbol)
    volume_24h = float(ticker_24h["quoteVolume"])

    return {
        "symbol":    symbol,
        "signals":   signals,
        "close":     last["close"],
        "ema20":     last["ema20"],
        "ema200":    last["ema200"],
        "vwap":      last["vwap"],
        "natr":      round(last["natr"], 3),
        "volText":   f"x{last['quote_volume']/avg_vol:.2f}",
        "prevVolCount": int((df.iloc[-5:-2]["quote_volume"] > last["quote_volume"]).sum()),
        "volume_24h": volume_24h,
        "swing_num": swing_num,
        "delta_pct": delta_pct,
        "relvol":    relvol if relvol is not None else "N/A",
        "atr_ratio_50": atr_ratio_50,
        "atr_ratio_75": atr_ratio_75,
        "atr_ratio_100": atr_ratio_100,
        "expansion_3": expansion_3,
        "expansion_5": expansion_5,
    }

# Глобальная переменная для markPrice WebSocket
_mark_twm = None
_mark_subscribed = set()
_mark_lock = Lock()

def _start_mark_twm():
    """Запускает _mark_twm в отдельном потоке с собственным event loop."""
    global _mark_twm
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    _mark_twm = ThreadedWebsocketManager()
    _mark_twm.start()

def subscribe_mark_price(symbol):
    global _mark_twm
    with _mark_lock:
        if symbol in _mark_subscribed:
            return
        try:
            if _mark_twm is None:
                t = Thread(target=_start_mark_twm, daemon=True)
                t.start()
                t.join(timeout=5)  # ждём пока запустится
            if _mark_twm is None:
                print(f"❌ markPrice: не удалось запустить TWM для {symbol}")
                return
            stream = f"{symbol.lower()}@markPrice@1s"
            _mark_twm.start_multiplex_socket(
                callback=handle_mark_price_global,
                streams=[stream]
            )
            _mark_subscribed.add(symbol)
            print(f"📡 markPrice подписка: {symbol}")
        except Exception as e:
            print(f"Ошибка подписки markPrice {symbol}: {e}")

def handle_mark_price_global(msg):
    try:
        print(f"🔍 markPrice raw: {str(msg)[:200]}")
        data = msg.get('data', msg)
        if data.get('e') != 'markPriceUpdate':
            return
        symbol = data.get('s')
        price  = float(data.get('p', 0))
        if price <= 0:
            return
        with TRADES_LOCK:
            has_open = any(
                trade["symbol"] == symbol and
                any(s["status"] == "OPEN" for s in trade["strategies"].values())
                for trade in ACTIVE_TRADES.values()
            )
        if not has_open:
            return
        check_and_close_strategies(symbol, price, price)
    except Exception as e:
        print(f"Ошибка handle_mark_price: {e}")

_need_restart = [False]

def sync_active_trades_with_db():
    import sqlite3
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trades.db")
    """Каждые 30 сек убирает из ACTIVE_TRADES стратегии закрытые вручную через дашборд."""
    while True:
        time.sleep(30)
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            with TRADES_LOCK:
                for trade_id, trade in list(ACTIVE_TRADES.items()):
                    rows = conn.execute(
                        "SELECT strategy, status FROM trade_strategies WHERE trade_id = ?",
                        (trade_id,)
                    ).fetchall()
                    for row in rows:
                        sname = row["strategy"]
                        if sname in trade["strategies"] and trade["strategies"][sname]["status"] == "OPEN":
                            if row["status"] not in ("OPEN", "SKIP"):
                                trade["strategies"][sname]["status"] = row["status"]
                                print(f"🔄 Синхронизировано: {trade_id} {sname} → {row['status']}")
                    # Если все закрыты — убираем из активных
                    if all(s["status"] != "OPEN" for s in trade["strategies"].values()):
                        del ACTIVE_TRADES[trade_id]
                        print(f"✅ Сделка {trade_id} удалена из активных")
            conn.close()
            save_active_trades()
        except Exception as e:
            print(f"Ошибка sync_active_trades: {e}")

# ================= MAIN =================
def main():
    global _mark_twm, _need_restart  # ← оба сюда
    init_db()
    # Переподписываемся на markPrice для уже открытых позиций при старте
    for trade in ACTIVE_TRADES.values():
        if any(s["status"] == "OPEN" for s in trade["strategies"].values()):
            subscribe_mark_price(trade["symbol"])
    Thread(target=sync_active_trades_with_db, daemon=True).start()
    symbols = get_liquid_futures_symbols()
    print(f"✅ Ликвидные токены: {len(symbols)}")

    last_signal_time = {}
    cooldown_seconds = COOLDOWN_BARS * 60 * 60  # кулдаун в часах

    def update_symbols_periodically():
        nonlocal symbols
        while True:
            time.sleep(3600)
            try:
                liquid = get_liquid_futures_symbols()
                open_syms = get_symbols_with_open_trades()
                symbols = list(set(liquid) | open_syms)
                print(f"♻️ Обновление: {len(liquid)} ликвидных + {len(open_syms)} с открытыми позициями")
            except Exception as e:
                print(f"Ошибка обновления токенов: {e}")

    Thread(target=update_symbols_periodically, daemon=True).start()

    task_queue = Queue()

    def process_signal(msg):
        # Быстрая проверка дня — до любых вычислений
        today = datetime.now(UTC).strftime("%A").lower()

        if today in SKIP_DAYS:
            return
        try:
            if msg.get("e") == "error":
                err_msg = msg.get("m","неизвестно")
                print(f"🔴 WebSocket ошибка: {msg}")
                if "reset" in str(err_msg).lower() or "closed" in str(err_msg).lower():
                    send_telegram(f"🔴 {BOT_NAME} WebSocket ошибка: {err_msg}")
                    _need_restart[0] = True
                return

            if 'data' not in msg or 'k' not in msg['data']:
                return
            candle = msg['data']['k']
            symbol = candle['s']
            if not candle['x']:
                return

            price_high = float(candle["h"])
            price_low  = float(candle["l"])

            # ===== Закрытие открытых стратегий =====
            check_and_close_strategies(symbol, price_high, price_low)

            # Новые сигналы только для символов из основного списка
            if symbol not in symbols:
                return

            # Cooldown
            now = time.time()
            if now - last_signal_time.get(symbol, 0) < cooldown_seconds:
                return

            # ===== Новые сигналы =====
            res = check_volume_signal(symbol)
            if not res:
                return

            last_signal_time[symbol] = now

            entry_price = res["close"]
            side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"

            # ===== Корреляция BTC =====
            try:
                btc_returns = get_btc_returns()
                if btc_returns is not None:
                    klines_sym = client.futures_klines(
                        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
                    )
                    df_sym = pd.DataFrame(klines_sym, columns=[
                        "open_time","open","high","low","close","volume",
                        "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
                    ])
                    df_sym["close"] = df_sym["close"].astype(float)
                    symbol_returns = df_sym["close"].pct_change()
                    btc_subset = btc_returns[-len(symbol_returns):]
                    corr = btc_subset.corr(symbol_returns)
                    corr_text = round(float(corr), 2) if corr is not None else "N/A"
                else:
                    corr_text = "N/A"
            except Exception as e:
                print(f"Ошибка корреляции {symbol}: {e}")
                corr_text = "N/A"

            # ===== Фильтр по корреляции =====
            if USE_CORREL_FILTER and isinstance(corr_text, float):
                side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"
                if side == "BUY":
                    if not apply_range_filter(corr_text, CORREL_MIN_BUY, CORREL_MAX_BUY):
                        print(f"⏭️ {symbol} пропущен — корреляция {corr_text} вне диапазона BUY")
                        return
                else:
                    if not apply_range_filter(corr_text, CORREL_MIN_SELL, CORREL_MAX_SELL):
                        print(f"⏭️ {symbol} пропущен — корреляция {corr_text} вне диапазона SELL")
                        return

            trade_id   = get_next_trade_id()
            # Стало — каждая стратегия проверяет свои фильтры:
            natr_val     = res["natr"]
            delta_val    = res.get("delta_pct", 0)
            vol_m        = res["volume_24h"] / 1_000_000  # или candle vol если нужен
            corr_val     = corr_text

            # Стало:
            strategies = {}
            for name, strat_cfg in STRATEGIES.items():
                if not check_strategy_filters(strat_cfg, side, natr_val, delta_val, vol_m, corr_val):
                    print(f"⏭️ {name} пропущена для {symbol} {side} — не прошла фильтры")
                    # Записываем прочерк — следим за ценой не будем, но в истории видно
                    skip_reasons = []
                    f = strat_cfg.get(side, {})
                    if f.get("USE_NATR_FILTER") and not apply_range_filter(res["natr"], f.get("NATR_MIN",0), f.get("NATR_MAX",0)):
                        skip_reasons.append("N")
                    if f.get("USE_DELTA_FILTER") and not apply_range_filter(delta_val if side=="BUY" else -delta_val, f.get("DELTA_MIN",0), f.get("DELTA_MAX",0)):
                        skip_reasons.append("D")
                    if f.get("USE_VOLUME_FILTER") and not apply_range_filter(vol_m, f.get("VOLUME_MIN",0), f.get("VOLUME_MAX",0)):
                        skip_reasons.append("V")
                    if f.get("USE_CORREL_FILTER"):
                        try:
                            if not apply_range_filter(float(corr_val), f.get("CORREL_MIN",0), f.get("CORREL_MAX",0)):
                                skip_reasons.append("C")
                        except (ValueError, TypeError):
                            pass
                    skip_str = "SKIP:" + "".join(skip_reasons) if skip_reasons else "SKIP:DAY"
                    strategies[name] = {"tp": None, "sl": None, "status": skip_str}
                    continue
                if side == "BUY":
                    tp = entry_price * (1 + strat_cfg["tp"])
                    sl = entry_price * (1 - abs(strat_cfg["sl"]))
                else:
                    tp = entry_price * (1 - strat_cfg["tp"])
                    sl = entry_price * (1 + abs(strat_cfg["sl"]))
                strategies[name] = {"tp": tp, "sl": sl, "status": "OPEN"}

            # Закрываем только если вообще ни одной OPEN стратегии
            if not any(s["status"] == "OPEN" for s in strategies.values()):
                print(f"⏭️ {symbol} {side} — все стратегии отфильтрованы, сделка не открыта")
                return

            with TRADES_LOCK:
                ACTIVE_TRADES[trade_id] = {
                    "symbol":      symbol,
                    "side":        side,
                    "entry_price": entry_price,
                    "strategies":  strategies,
                    "open_time":   datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
                }
            save_active_trades()

            # ===== Подписываемся на markPrice для нового символа =====
            subscribe_mark_price(symbol)

            write_trade_to_excel(
                trade_id,
                {
                    "symbol":      symbol,
                    "signals":     res["signals"],
                    "strategies":  strategies,
                    "entry_price": entry_price,
                    "natr":        res["natr"],
                    "swing_num":   res["swing_num"],
                    "delta_pct":   res["delta_pct"],
                    "relvol":      res["relvol"],
                    "atr_ratio_50": res["atr_ratio_50"],
                    "atr_ratio_75": res["atr_ratio_75"],
                    "atr_ratio_100": res["atr_ratio_100"],
                    "expansion_3": res["expansion_3"],
                    "expansion_5": res["expansion_5"],
                },
                vol_text=res["volText"],
                vol24=res["volume_24h"] / 1_000_000,
                corr_text=corr_text
            )
            
            insert_trade(trade_id, BOT_NAME, {   # ← добавить
                "symbol": symbol,
                "side": side,
                "entry_price": entry_price,
                "signals": res["signals"],
                "strategies": strategies,
                "natr": res["natr"],
                "swing_num": res["swing_num"],
                "delta_pct": res.get("delta_pct"),
                "relvol": res.get("relvol"),
                "atr_ratio_50": res["atr_ratio_50"],
                "atr_ratio_75": res["atr_ratio_75"],
                "atr_ratio_100": res["atr_ratio_100"],
                "expansion_3": res["expansion_3"],
                "expansion_5": res["expansion_5"],
            }, vol_text=res["volText"], vol_24h=res["volume_24h"]/1_000_000, corr_text=corr_text)

            # ===== Telegram =====
            vol24 = res["volume_24h"] / 1_000_000
            msg_text = (
                f"🤖 {BOT_NAME}\n"
                f"🔥 {res['symbol']}\n"
                f"Тип: {', '.join(res['signals'])}\n"
                f"Close: {res['close']:.6f}\n"
                f"EMA20: {res['ema20']:.6f}\n"
                f"EMA200: {res['ema200']:.6f}\n"
                f"VWAP: {res['vwap']:.6f}\n"
                f"VOL {res['volText']}\n"
                f"Prev volume higher: {res['prevVolCount']}/3\n"
                f"VOL 24h: {vol24:.1f}M USDT\n"
                f"Corr BTC: {corr_text}\n"
                f"NATR: {res['natr']}%\n"
                f"Свинг: {res['swing_num']}\n"
                f"Delta: {res['delta_pct']}%\n"
                f"RelVol: {res['relvol']}\n"
            )
            print(msg_text)

        except Exception as e:
            print(f"Ошибка process_signal: {e}")

    def handle_kline(msg):
        task_queue.put(msg)

    def worker():
        while True:
            msg = task_queue.get()
            process_signal(msg)
            task_queue.task_done()

    Thread(target=worker, daemon=True).start()

    # ===== WebSocket с переподключением и плановым перезапуском =====
    chunk_size = 10

    while True:
        try:
            open_syms = get_symbols_with_open_trades()
            all_symbols = list(set(symbols) | open_syms)

            twm = ThreadedWebsocketManager()
            twm.start()

            for i in range(0, len(all_symbols), chunk_size):
                streams = [f"{s.lower()}@kline_1h" for s in all_symbols[i:i+chunk_size]]
                twm.start_multiplex_socket(callback=handle_kline, streams=streams)

            print("🟢 WebSocket запущен")
            send_telegram(f"🟢 {BOT_NAME} WebSocket запущен")

            for _ in range(24 * 60):
                time.sleep(60)
                if _need_restart[0]:
                    print("🔄 Принудительный перезапуск по флагу ошибки...")
                    _need_restart[0] = False
                    break

            print("♻️ Перезапуск WebSocket...")
            save_active_trades()
            with _mark_lock:
                _mark_subscribed.clear()
                if _mark_twm is not None:
                    try:
                        _mark_twm.stop()
                    except Exception:
                        pass
                    _mark_twm = None
            twm.stop()

        except Exception as e:
            err_str = str(e)
            print(f"🔴 WebSocket упал: {err_str}. Переподключение через 30 секунд...")
            if "Socket Manager" not in err_str and "initialize" not in err_str:
                send_telegram(f"🔴 {BOT_NAME} WebSocket упал: {err_str}")
            save_active_trades()
            try:
                twm.stop()          # ← добавить
            except Exception:
                pass
            time.sleep(30)


if __name__ == "__main__":
    main()
