from binance.client import Client
from binance import ThreadedWebsocketManager
import pandas as pd
import time
from datetime import datetime, timezone
import requests
import os
import json
import argparse
import openpyxl
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from threading import Thread, Lock
from queue import Queue

# ===== ЗАГРУЗКА КОНФИГА =====
parser = argparse.ArgumentParser()
parser.add_argument("--config", required=True)
args = parser.parse_args()

with open(args.config, "r") as f:
    config = json.load(f)

BOT_NAME = config["NAME"]

load_dotenv()

# ================= НАСТРОЙКИ =================
MIN_24H_VOLUME   = config["MIN_24H_VOLUME"]
LOOKBACK_CANDLES = config["LOOKBACK_CANDLES"]
VOLUME_LOOKBACK  = config["VOLUME_LOOKBACK"]

VOL_MULT         = float(config["VOL_MULT"])
MIN_BODY_PCT     = float(config["MIN_BODY_PCT"])
COOLDOWN_BARS    = config["COOLDOWN_BARS"]

EMA_FAST         = config["EMA_FAST"]
EMA_SLOW         = config["EMA_SLOW"]

BTC_LOOKBACK     = config["BTC_LOOKBACK"]
ATR_LEN          = config["ATR_LEN"]

# Фильтры — включить/выключить
USE_EMA_FILTER  = config.get("USE_EMA_FILTER", True)
USE_VWAP_FILTER = config.get("USE_VWAP_FILTER", True)

EXCEL_STRAT_START_COL = 14  # колонка N
PREV_VOL_WINDOW  = 3

CHAT_ID   = os.getenv("CHAT_ID")
BOT_TOKEN = os.getenv("BOT_TOKEN")

client = Client()
BLACKLIST = {
    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT",
    "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT"
}

# ================= TRADES =================
TRADE_STATE_FILE   = f"trades_state_{BOT_NAME}.json"
EXCEL_FILE         = f"trades_{BOT_NAME}.xlsx"
ACTIVE_TRADES_FILE = f"active_trades_{BOT_NAME}.json"

TRADES_LOCK = Lock()
EXCEL_LOCK  = Lock()
_ID_LOCK    = Lock()

SHEET_MAP = {
    "CONFSP1": "confsp1",
    "CONFSP2": "confsp2",
    "CONFSP3": "confsp3",
    "CONFSP4": "confsp4",
}

# Стратегии: 3:1, 6:1, 6:2, 10:3, 12:4
STRATEGIES = {
    "3:1":  {"tp": 0.03,  "sl": -0.01},
    "6:1":  {"tp": 0.06,  "sl": -0.01},
    "6:2":  {"tp": 0.06,  "sl": -0.02},
    "10:3": {"tp": 0.10,  "sl": -0.03},
    "12:4": {"tp": 0.12,  "sl": -0.04},
}

def load_trade_id():
    if not os.path.exists(TRADE_STATE_FILE):
        return 0
    with open(TRADE_STATE_FILE, "r") as f:
        return json.load(f).get("last_trade_id", 0)

def save_trade_id(tid):
    with open(TRADE_STATE_FILE, "w") as f:
        json.dump({"last_trade_id": tid}, f)

def save_active_trades():
    with TRADES_LOCK:
        with open(ACTIVE_TRADES_FILE, "w") as f:
            json.dump(ACTIVE_TRADES, f)

def load_active_trades():
    if not os.path.exists(ACTIVE_TRADES_FILE):
        return {}
    with open(ACTIVE_TRADES_FILE, "r") as f:
        return json.load(f)

ACTIVE_TRADES = load_active_trades()
LAST_TRADE_ID = load_trade_id()

def get_next_trade_id():
    global LAST_TRADE_ID
    with _ID_LOCK:
        LAST_TRADE_ID += 1
        save_trade_id(LAST_TRADE_ID)
        return f"{LAST_TRADE_ID:05d}"

# ================= TELEGRAM =================
def send_telegram(message: str):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": CHAT_ID, "text": message}
    try:
        requests.post(url, data=payload, timeout=10)
    except Exception as e:
        print(f"Ошибка Telegram: {e}")

# ================= EXCEL =================
def write_trade_to_excel(trade_id, trade_info, vol_text, vol24, corr_text):
    sheet_name = SHEET_MAP.get(BOT_NAME, "confsp1")

    with EXCEL_LOCK:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            for sn in SHEET_MAP.values():
                if sn not in wb.sheetnames:
                    wb.create_sheet(sn)
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            wb.save(EXCEL_FILE)

        wb = openpyxl.load_workbook(EXCEL_FILE)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        headers = {
            "A":"Дата","B":"Время","C":"День","D":"Тикет","E":"Объем",
            "F":"Trade_id","G":"Тип","H":"Импульс","J":"Цена входа",
            "K":"Корреляция","M":"NATR%",
            "N":"3:1","O":"6:1","P":"6:2","Q":"10:3","R":"12:4",
            "S":"3:1 цена","T":"6:1 цена","U":"6:2 цена",
            "V":"10:3 цена","W":"12:4 цена",
        }
        if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
            for col, header in headers.items():
                ws[f"{col}1"] = header

        next_row = ws.max_row + 1
        dt = datetime.now()
        ws["A"+str(next_row)] = dt.strftime("%d.%m.%Y")
        ws["B"+str(next_row)] = dt.strftime("%H:%M:%S")
        ws["C"+str(next_row)] = dt.strftime("%a")
        ws["D"+str(next_row)] = trade_info["symbol"]
        ws["E"+str(next_row)] = vol24
        ws["F"+str(next_row)] = trade_id
        ws["G"+str(next_row)] = ", ".join(trade_info["signals"])
        ws["H"+str(next_row)] = vol_text
        ws["J"+str(next_row)] = trade_info["entry_price"]
        ws["K"+str(next_row)] = corr_text
        ws["M"+str(next_row)] = trade_info["natr"]

        for idx, s in enumerate(STRATEGIES.keys()):
            col = get_column_letter(EXCEL_STRAT_START_COL + idx)
            ws[f"{col}{next_row}"] = trade_info["strategies"][s]["status"]

        wb.save(EXCEL_FILE)

def update_trade_status_in_excel(trade_id, strategy_name, status, close_price):
    sheet_name = SHEET_MAP.get(BOT_NAME, "confsp1")

    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]

        for row in range(2, ws.max_row+1):
            if str(ws[f"F{row}"].value) == trade_id:
                col_map_status  = {"3:1":"N","6:1":"O","6:2":"P","10:3":"Q","12:4":"R"}
                col_map_details = {"3:1":"S","6:1":"T","6:2":"U","10:3":"V","12:4":"W"}
                col_s = col_map_status[strategy_name]
                col_d = col_map_details[strategy_name]
                ws[f"{col_s}{row}"] = status
                ws[f"{col_d}{row}"] = round(close_price, 6)
                break

        wb.save(EXCEL_FILE)

# ================= INDICATORS =================
def calculate_session_vwap(df):
    df = df.copy()
    df["date"] = pd.to_datetime(df["open_time"], unit="ms").dt.date
    tp = (df["high"] + df["low"] + df["close"]) / 3
    df["tpv"] = tp * df["volume"]
    df["cum_tpv"] = df.groupby("date")["tpv"].cumsum()
    df["cum_vol"] = df.groupby("date")["volume"].cumsum()
    return df["cum_tpv"] / df["cum_vol"]

def calculate_atr(df, period):
    hl = df["high"] - df["low"]
    hc = (df["high"] - df["close"].shift()).abs()
    lc = (df["low"] - df["close"].shift()).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def get_liquid_futures_symbols():
    tickers = client._request_futures_api(method="get", path="ticker/24hr")
    symbols = []
    for t in tickers:
        symbol = t["symbol"]
        if not symbol.endswith("USDT") or symbol in BLACKLIST:
            continue
        if float(t["quoteVolume"]) < MIN_24H_VOLUME:
            continue
        symbols.append(symbol)
    return symbols

def get_btc_returns():
    try:
        klines_btc = client.futures_klines(
            symbol="BTCUSDT", interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
        )
        df_btc = pd.DataFrame(klines_btc, columns=[
            "open_time","open","high","low","close","volume",
            "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
        ])
        df_btc["close"] = df_btc["close"].astype(float)
        return df_btc["close"].pct_change()
    except Exception as e:
        print(f"Ошибка загрузки BTC свечей: {e}")
        return None

def check_volume_signal(symbol):
    klines = client.futures_klines(
        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=LOOKBACK_CANDLES
    )
    df = pd.DataFrame(klines, columns=[
        "open_time","open","high","low","close",
        "volume","close_time","quote_volume",
        "trades","taker_buy_base","taker_buy_quote","ignore"
    ])
    for c in ["open","high","low","close","volume"]:
        df[c] = df[c].astype(float)

    df["ema20"]  = df["close"].ewm(span=EMA_FAST, adjust=False).mean()
    df["ema200"] = df["close"].ewm(span=EMA_SLOW, adjust=False).mean()
    df["atr"]    = calculate_atr(df, ATR_LEN)
    df["natr"]   = (df["atr"] / df["close"]) * 100
    df["vwap"]   = calculate_session_vwap(df)
    df["quote_volume"] = df["close"] * df["volume"]

    avg_vol = df["quote_volume"].iloc[-(VOLUME_LOOKBACK + 2):-2].mean()
    last = df.iloc[-2]

    volume_spike = last["quote_volume"] >= avg_vol * VOL_MULT

    body     = abs(last["close"] - last["open"])
    rng      = last["high"] - last["low"]
    body_pct = 0 if rng == 0 else body / rng * 100
    bull = last["close"] > last["open"]
    bear = last["close"] < last["open"]

    strong_body = body_pct >= MIN_BODY_PCT

    # EMA фильтр
    bull_trend = last["ema20"] > last["ema200"]
    bear_trend = last["ema20"] < last["ema200"]
    ema_bull_ok = bull_trend if USE_EMA_FILTER else True
    ema_bear_ok = bear_trend if USE_EMA_FILTER else True

    # VWAP фильтр
    below_vwap = (last["close"] < last["vwap"]) if USE_VWAP_FILTER else True
    above_vwap = (last["close"] > last["vwap"]) if USE_VWAP_FILTER else True

    # Cooldown
    if COOLDOWN_BARS > 0:
        recent = df.iloc[-(COOLDOWN_BARS + 2):-2]
        recent_spike = (recent["quote_volume"] >= avg_vol * VOL_MULT).any()
    else:
        recent_spike = False

    signals = []
    if volume_spike and bull and strong_body and ema_bull_ok and below_vwap and not recent_spike:
        signals.append("BUY")
    if volume_spike and bear and strong_body and ema_bear_ok and above_vwap and not recent_spike:
        signals.append("SELL")

    if not signals:
        return None

    ticker_24h = client.futures_ticker(symbol=symbol)
    volume_24h = float(ticker_24h["quoteVolume"])

    return {
        "symbol":    symbol,
        "signals":   signals,
        "close":     last["close"],
        "ema20":     last["ema20"],
        "ema200":    last["ema200"],
        "vwap":      last["vwap"],
        "natr":      round(last["natr"], 3),
        "volText":   f"x{last['quote_volume']/avg_vol:.2f}",
        "prevVolCount": int((df.iloc[-5:-2]["quote_volume"] > last["quote_volume"]).sum()),
        "volume_24h": volume_24h,
    }

# ================= MAIN =================
def main():
    symbols = get_liquid_futures_symbols()
    print(f"✅ Ликвидные токены: {len(symbols)}")

    last_signal_time = {}
    cooldown_seconds = COOLDOWN_BARS * 60 * 60  # кулдаун в часах

    def update_symbols_periodically():
        nonlocal symbols
        while True:
            time.sleep(3600)
            try:
                symbols = get_liquid_futures_symbols()
                print(f"♻️ Обновление токенов: {len(symbols)}")
            except Exception as e:
                print(f"Ошибка обновления токенов: {e}")

    Thread(target=update_symbols_periodically, daemon=True).start()

    task_queue = Queue()

    def process_signal(msg):
        try:
            if msg.get("e") == "error":
                print(f"🔴 WebSocket ошибка: {msg}")
                send_telegram(f"🔴 {BOT_NAME} WebSocket ошибка: {msg.get('m', 'неизвестно')}")
                return

            if 'data' not in msg or 'k' not in msg['data']:
                return
            candle = msg['data']['k']
            symbol = candle['s']
            if symbol not in symbols or not candle['x']:
                return

            price_high = float(candle["h"])
            price_low  = float(candle["l"])

            # ===== Закрытие открытых стратегий =====
            closed_trades = []
            with TRADES_LOCK:
                for trade_id, trade in list(ACTIVE_TRADES.items()):
                    if trade["symbol"] != symbol:
                        continue
                    for strat_name, strat in trade["strategies"].items():
                        if strat["status"] != "OPEN":
                            continue
                        if trade["side"] == "BUY":
                            if price_low <= strat["sl"]:
                                result = "SL"
                            elif price_high >= strat["tp"]:
                                result = "TP"
                            else:
                                continue
                        else:
                            if price_high >= strat["sl"]:
                                result = "SL"
                            elif price_low <= strat["tp"]:
                                result = "TP"
                            else:
                                continue

                        strat["status"] = result
                        close_price = strat["sl"] if result == "SL" else strat["tp"]
                        Thread(target=update_trade_status_in_excel,
                               args=(trade_id, strat_name, result, close_price), daemon=True).start()

                    if all(s["status"] != "OPEN" for s in trade["strategies"].values()):
                        closed_trades.append(trade_id)

                for tid in closed_trades:
                    del ACTIVE_TRADES[tid]

            if closed_trades:
                save_active_trades()

            # Cooldown
            now = time.time()
            if now - last_signal_time.get(symbol, 0) < cooldown_seconds:
                return

            # ===== Новые сигналы =====
            res = check_volume_signal(symbol)
            if not res:
                return

            last_signal_time[symbol] = now

            entry_price = res["close"]
            side = "BUY" if any("BUY" in s for s in res["signals"]) else "SELL"

            # ===== Корреляция BTC =====
            try:
                btc_returns = get_btc_returns()
                if btc_returns is not None:
                    klines_sym = client.futures_klines(
                        symbol=symbol, interval=Client.KLINE_INTERVAL_1HOUR, limit=BTC_LOOKBACK
                    )
                    df_sym = pd.DataFrame(klines_sym, columns=[
                        "open_time","open","high","low","close","volume",
                        "close_time","quote_volume","trades","taker_buy_base","taker_buy_quote","ignore"
                    ])
                    df_sym["close"] = df_sym["close"].astype(float)
                    symbol_returns = df_sym["close"].pct_change()
                    btc_subset = btc_returns[-len(symbol_returns):]
                    corr = btc_subset.corr(symbol_returns)
                    corr_text = round(float(corr), 2) if corr is not None else "N/A"
                else:
                    corr_text = "N/A"
            except Exception as e:
                print(f"Ошибка корреляции {symbol}: {e}")
                corr_text = "N/A"

            trade_id   = get_next_trade_id()
            strategies = {}
            for name, strat_cfg in STRATEGIES.items():
                if side == "BUY":
                    tp = entry_price * (1 + strat_cfg["tp"])
                    sl = entry_price * (1 - abs(strat_cfg["sl"]))
                else:
                    tp = entry_price * (1 - strat_cfg["tp"])
                    sl = entry_price * (1 + abs(strat_cfg["sl"]))
                strategies[name] = {"tp": tp, "sl": sl, "status": "OPEN"}

            with TRADES_LOCK:
                ACTIVE_TRADES[trade_id] = {
                    "symbol":      symbol,
                    "side":        side,
                    "entry_price": entry_price,
                    "strategies":  strategies,
                    "open_time":   datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                }
            save_active_trades()

            write_trade_to_excel(
                trade_id,
                {
                    "symbol":      symbol,
                    "signals":     res["signals"],
                    "strategies":  strategies,
                    "entry_price": entry_price,
                    "natr":        res["natr"],
                },
                vol_text=res["volText"],
                vol24=res["volume_24h"] / 1_000_000,
                corr_text=corr_text
            )

            # ===== Telegram =====
            vol24 = res["volume_24h"] / 1_000_000
            msg_text = (
                f"🤖 {BOT_NAME}\n"
                f"🔥 {res['symbol']}\n"
                f"Тип: {', '.join(res['signals'])}\n"
                f"Close: {res['close']:.6f}\n"
                f"EMA20: {res['ema20']:.6f}\n"
                f"EMA200: {res['ema200']:.6f}\n"
                f"VWAP: {res['vwap']:.6f}\n"
                f"VOL {res['volText']}\n"
                f"Prev volume higher: {res['prevVolCount']}/3\n"
                f"VOL 24h: {vol24:.1f}M USDT\n"
                f"Corr BTC: {corr_text}\n"
                f"NATR: {res['natr']}%\n"
            )
            print(msg_text)
            send_telegram(msg_text)

        except Exception as e:
            print(f"Ошибка process_signal: {e}")

    def handle_kline(msg):
        task_queue.put(msg)

    def worker():
        while True:
            msg = task_queue.get()
            process_signal(msg)
            task_queue.task_done()

    Thread(target=worker, daemon=True).start()

    # ===== WebSocket с переподключением и плановым перезапуском =====
    chunk_size = 30

    while True:
        try:
            twm = ThreadedWebsocketManager()
            twm.start()

            for i in range(0, len(symbols), chunk_size):
                streams = [f"{s.lower()}@kline_1h" for s in symbols[i:i+chunk_size]]
                twm.start_multiplex_socket(callback=handle_kline, streams=streams)

            print("🟢 WebSocket запущен")
            send_telegram(f"🟢 {BOT_NAME} WebSocket запущен")

            time.sleep(24 * 60 * 60)
            print("♻️ Плановый перезапуск WebSocket...")
            send_telegram(f"♻️ {BOT_NAME} плановый перезапуск WebSocket")
            save_active_trades()
            twm.stop()

        except Exception as e:
            print(f"🔴 WebSocket упал: {e}. Переподключение через 30 секунд...")
            send_telegram(f"🔴 {BOT_NAME} WebSocket упал: {e}. Переподключение через 30 секунд...")
            save_active_trades()
            try:
                twm.stop()
            except Exception:
                pass
            time.sleep(30)


if __name__ == "__main__":
    main()
